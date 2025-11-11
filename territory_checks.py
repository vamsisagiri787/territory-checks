from __future__ import annotations
# ==============================================================================
#  Author : Vamsi Krishna S. (enterprise build)
#  Program: Territory Checks – Weekly Brand × Broker Counter (Sun→Sat)
#  Build  : 2.5 (enterprise, 2025-11-10)
#
#  WHAT'S IN THIS BUILD
#  --------------------
#  • Week window = previous Sun 00:00:00 → this Sun 00:00:00 (UTC, exclusive)
#  • Brand detection prefers To/Cc/Bcc, then routing headers only (no noise)
#  • Growth Coach alias includes growthcoach.com
#  • Broker detection: subject → forwarded body header → sender domain
#  • Territory: subject → bodyPreview → full HTML body (fallback)
#  • De-dup: Brand × Broker × Territory (keeps first)
#  • Always creates per-brand sheets even if weekly count is 0
#  • Unmapped + Audit tabs with helpful columns
#  • No secrets in code (all from env)
# ==============================================================================

# ================================= Imports ====================================
import os, re, time, json, logging, unicodedata, html
from datetime import datetime, timedelta, timezone
from typing import Dict, Iterable, List, Tuple, Optional

import requests, msal, pandas as pd
from openpyxl import Workbook, load_workbook

# =============================== Configuration ================================
TENANT_ID      = os.getenv("GRAPH_TENANT_ID", "")
CLIENT_ID      = os.getenv("GRAPH_CLIENT_ID", "")
CLIENT_SECRET  = os.getenv("GRAPH_CLIENT_SECRET", "")
MASTER_MAILBOX = os.getenv("MASTER_MAILBOX", "territorycheck@strategicfranchising.com")
OUT_DIR        = os.getenv("OUT_DIR", "/tmp")
LOG_LEVEL      = os.getenv("LOG_LEVEL", "INFO").upper().strip()

COUNT_FORWARDS = True
SKIP_REPLIES   = True
SHOW_SUBJECT_IN_DETAILS = True

USER_AGENT   = "TerritoryChecks/2.5"
HTTP_TIMEOUT = 60
MAX_RETRIES  = 5

logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="%(asctime)s | %(levelname)-8s | %(message)s",
)
log = logging.getLogger("territory-checks")

# ================================ Brand Rules =================================
BRAND_ADDRESSES: Dict[str, List[str]] = {
    "Caring Transitions": ["territorycheck@caringtransitions.com", "caringtransitions.com"],
    "Fresh Coat"        : ["territorycheck@freshcoatpainters.com", "freshcoatpainters.com", "freshcoat.com"],
    "TruBlue"           : ["territorycheck@trublueally.com", "trublueally.com"],
    "Growth Coach"      : ["territorycheck@thegrowthcoach.com", "thegrowthcoach.com", "growthcoach.com"],
    "Pet Wants"         : ["territorycheck@petwants.com", "petwants.com"],
}

# ============================== Broker Mapping ================================
BROKER_SUBJECT_KEYWORDS: Dict[str, str] = {
    "FranServe"        : r"\bfran\s*serve|franservesupport|franserve\b",
    "IFPG"             : r"\bifpg\b",
    "BAI"              : r"\bb\.?a\.?i\b|business\W*alliance",
    "TES"              : r"\btes\b|the\s*entrepreneur.?s\s*source|franchisesource|esourcecoach",
    "FranNet"          : r"\bfrannet\b",
    "FBA"              : r"\bfba\b|franchiseba|franchiseiba|fbamembers",
    "TPF"              : r"\btpf\b|the\s*perfect\s*franchise",
    "FCC"              : r"\bfcc\b|franchise\s*consulting\s*company",
    "Franchise Empire" : r"\bfranchise\s*empire\b",
    "SFA Advisors"     : r"\bsfa\s*advisors\b|\bsuccess\s*franchise\s*advisors\b|\bsuccess\s*fran\b",
}
BROKER_DOMAIN_MAP: Dict[str, str] = {
    "franserve.com":"FranServe","franservesupport.com":"FranServe","focusonfranchising.com":"FranServe",
    "ifpg.org":"IFPG","ifpg.com":"IFPG",
    "businessallianceinc.com":"BAI","zizefranchise.com":"BAI","inspirefranchiseconsulting.com":"BAI",
    "franchise-connector.com":"BAI","franchiseconnector.com":"BAI","markfranchise.com":"BAI",
    "franchisesource.com":"TES","esourcecoach.com":"TES",
    "frannet.com":"FranNet",
    "franchiseba.com":"FBA","franchiseiba.com":"FBA","fbamembers.com":"FBA",
    "theperfectfranchise.com":"TPF",
    "thefranchiseconsultingcompany.com":"FCC",
    "franchiseempire.com":"Franchise Empire",
    "sfaadvisors.com":"SFA Advisors","successfran.net":"SFA Advisors",
    "securefranchise.com":"SFA Advisors","successfranchiseadvisors.com":"SFA Advisors",
}
BROKER_ORDER = ["IFPG","FranServe","BAI","TES","FranNet","FBA","TPF","FCC","Franchise Empire","SFA Advisors","Others"]

# ============================== Date/Time Window ==============================
def last_completed_week_utc(now_utc: Optional[datetime] = None) -> Tuple[datetime, datetime, datetime, str, str]:
    """
    Last fully completed week (UTC):
      start (inclusive)  = previous Sun 00:00:00
      end_exclusive      = this    Sun 00:00:00
      end_inclusive      = this    Sat 23:59:59 (labels only)
    """
    now = now_utc or datetime.now(timezone.utc).replace(microsecond=0)
    days_since_sun = (now.weekday() + 1) % 7
    this_sun = (now - timedelta(days=days_since_sun)).replace(hour=0, minute=0, second=0, microsecond=0)
    prior_sun = this_sun - timedelta(days=7)

    start = prior_sun
    end_exclusive = this_sun
    end_inclusive = this_sun - timedelta(seconds=1)

    week_label = f"{start:%m.%d}-{end_inclusive:%m.%d.%y}"
    file_label = f"{end_inclusive:%Y-%m-%d}"
    return start, end_inclusive, end_exclusive, week_label, file_label

# ================================ Graph Client ================================
class GraphClient:
    authority = "https://login.microsoftonline.com/{tenant}"
    scope     = ["https://graph.microsoft.com/.default"]
    base      = "https://graph.microsoft.com/v1.0"

    def __init__(self, tenant_id: str, client_id: str, client_secret: str, mailbox: str):
        self.tenant_id, self.client_id, self.client_secret, self.mailbox = tenant_id, client_id, client_secret, mailbox
        self._token: Optional[str] = None

    def authenticate(self) -> str:
        app = msal.ConfidentialClientApplication(
            self.client_id,
            authority=self.authority.format(tenant=self.tenant_id),
            client_credential=self.client_secret,
        )
        result = app.acquire_token_for_client(scopes=self.scope)
        if "access_token" not in result:
            raise SystemExit("Graph token error:\n" + json.dumps(result, indent=2))
        self._token = result["access_token"]
        return self._token

    def _headers(self) -> dict:
        if not self._token:
            self.authenticate()
        return {"Authorization": f"Bearer {self._token}", "User-Agent": USER_AGENT}

    def _get(self, url: str, params: dict | None = None) -> requests.Response:
        headers = self._headers()
        backoff = 2
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                r = requests.get(url, headers=headers, params=params, timeout=HTTP_TIMEOUT)
            except requests.RequestException as e:
                log.warning("Network error (%d/%d): %s", attempt, MAX_RETRIES, e)
                time.sleep(backoff); backoff = min(backoff*2, 32); continue
            if r.status_code in (429,500,502,503,504):
                wait = int(r.headers.get("Retry-After", backoff))
                log.warning("Graph %s — waiting %ss (%d/%d)", r.status_code, wait, attempt, MAX_RETRIES)
                time.sleep(wait); backoff = min(backoff*2,32); continue
            if r.status_code == 401:
                log.warning("401 Unauthorized — refreshing token")
                self.authenticate(); headers = self._headers(); continue
            if r.status_code == 403:
                raise SystemExit("403 Forbidden. Check App Access Policy.")
            r.raise_for_status(); return r
        r.raise_for_status(); return r

    def _list_child_folders(self, parent_id: str | None = None) -> List[dict]:
        url = f"{self.base}/users/{self.mailbox}/mailFolders" + (f"/{parent_id}/childFolders" if parent_id else "")
        params = {"$top": 50, "$select": "id,displayName,childFolderCount"}
        out: List[dict] = []
        while True:
            data = self._get(url, params=params).json()
            out.extend(data.get("value", []))
            nxt = data.get("@odata.nextLink")
            if not nxt: break
            url, params = nxt, None
        final, seen = [], set()
        for f in out:
            if f["id"] not in seen:
                final.append(f); seen.add(f["id"])
            if f.get("childFolderCount", 0) > 0:
                final.extend(self._list_child_folders(f["id"]))
        return final

    def list_all_folders(self) -> List[dict]:
        folders = self._list_child_folders(None)
        log.info("Discovered %d mail folders.", len(folders))
        return folders

    def fetch_messages_in_folder(self, folder_id: str, start_iso: str, end_iso: str) -> Iterable[dict]:
        url = f"{self.base}/users/{self.mailbox}/mailFolders/{folder_id}/messages"
        params = {
            "$filter": f"receivedDateTime ge {start_iso} and receivedDateTime lt {end_iso}",
            "$orderby": "receivedDateTime asc",
            "$top": 50,
            "$select": (
                "id,internetMessageId,conversationId,receivedDateTime,subject,"
                "from,toRecipients,ccRecipients,bccRecipients,bodyPreview,internetMessageHeaders"
            )
        }
        while True:
            data = self._get(url, params=params).json()
            for m in data.get("value", []):
                yield m
            nxt = data.get("@odata.nextLink")
            if not nxt: break
            url, params = nxt, None

    def fetch_full_body(self, message_id: str) -> str:
        url = f"{self.base}/users/{self.mailbox}/messages/{message_id}"
        params = {"$select": "body"}  # body.content (HTML)
        data = self._get(url, params=params).json()
        content = (((data.get("body") or {}).get("content")) or "")
        return content

# ============================ Classification Helpers ==========================
RE_PREFIX = re.compile(r"^\s*re\s*[:\-]\s*", re.IGNORECASE)
FW_PREFIX = re.compile(r"^\s*(fw|fwd)\s*[:\-]\s*", re.IGNORECASE)
def is_reply(s: str) -> bool:   return bool(RE_PREFIX.match(s or ""))
def is_forward(s: str) -> bool: return bool(FW_PREFIX.match(s or ""))

def _norm(text: str) -> str:
    s = unicodedata.normalize("NFKC", text or "").lower()
    s = re.sub(r"[\u00A0\u2000-\u200B\u2060]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _html_to_text(html_str: str) -> str:
    if not html_str: return ""
    txt = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", html_str)
    txt = re.sub(r"(?is)<br\s*/?>", "\n", txt)
    txt = re.sub(r"(?is)</p\s*>", "\n", txt)
    txt = re.sub(r"(?is)<.*?>", " ", txt)
    txt = html.unescape(txt)
    txt = re.sub(r"\r", "\n", txt)
    return re.sub(r"[ \t]+", " ", txt)

def _addr_list(*lists: List[dict] | None) -> List[str]:
    addrs: List[str] = []
    for coll in lists:
        for item in (coll or []):
            a = (item.get("emailAddress") or {}).get("address", "")
            if a: addrs.append(a.lower())
    return addrs

def _match_brand_in_text(haystack: str) -> Optional[str]:
    hay = haystack.lower()
    for brand, needles in BRAND_ADDRESSES.items():
        for n in needles:
            tok = n.lower()
            if "@" in tok:
                if re.search(r"\b" + re.escape(tok) + r"\b", hay):
                    return brand
            else:
                if re.search(r"(?<![a-z0-9])" + re.escape(tok) + r"\b", hay):
                    return brand
    return None

def brand_from_recipients(to_list: List[dict] | None,
                          cc_list: List[dict] | None,
                          bcc_list: List[dict] | None) -> tuple[Optional[str], str]:
    addrs = _addr_list(to_list, cc_list, bcc_list)
    hay = " ".join(addrs)
    b = _match_brand_in_text(hay)
    return b, ("To/Cc/Bcc" if b else "")

_ROUTING_HEADERS = {
    "to","cc","bcc","delivered-to","x-original-to","x-envelope-to","x-forwarded-to","return-path"
}
def brand_from_headers(headers: List[dict] | None) -> tuple[Optional[str], str]:
    if not headers: return (None, "")
    vals = []
    for h in headers:
        name = (h.get("name") or "").lower()
        if name in _ROUTING_HEADERS:
            vals.append(h.get("value",""))
    if not vals: return (None, "")
    joined = " ".join(vals)
    b = _match_brand_in_text(joined)
    return b, ("Headers(" + ",".join(sorted(_ROUTING_HEADERS)) + ")" if b else "")

def broker_from_subject(subject: str) -> Optional[str]:
    s = _norm(subject)
    for broker, pattern in BROKER_SUBJECT_KEYWORDS.items():
        if re.search(pattern, s):
            return broker
    return None

def broker_from_sender(from_obj: dict | None) -> str:
    addr = ((from_obj or {}).get("emailAddress") or {}).get("address", "") or ""
    dom  = addr.split("@")[-1].lower() if "@" in addr else addr.lower()
    for d, canonical in BROKER_DOMAIN_MAP.items():
        if d in dom:
            return canonical
    return "Others"

RE_FWD_FROM_ANGLE = re.compile(r"^from:\s*[^<\r\n]*<\s*([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})\s*>",
                               re.IGNORECASE | re.MULTILINE)
RE_FWD_FROM_BARE  = re.compile(r"^from:\s*([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})",
                               re.IGNORECASE | re.MULTILINE)
def forwarded_broker_from_bodypreview(body_preview: str) -> Optional[str]:
    if not body_preview: return None
    m = RE_FWD_FROM_ANGLE.search(body_preview) or RE_FWD_FROM_BARE.search(body_preview)
    if m:
        email = m.group(1).strip().lower()
        dom = email.split("@")[-1]
        for d, canonical in BROKER_DOMAIN_MAP.items():
            if d in dom:
                return canonical
    for broker, pattern in BROKER_SUBJECT_KEYWORDS.items():
        if re.search(pattern, (body_preview or "").lower()):
            return broker
    return None

# ---------- Territory matchers ----------
TERR_PATTERNS = (
    re.compile(r"\bin\s+([a-z .'\-]+,\s*[a-z]{2})\b", re.IGNORECASE),
    re.compile(r"\b([a-z .'\-]+,\s*[a-z]{2})\b", re.IGNORECASE),
)
BODY_TERR_PATTERNS = (
    re.compile(r"desired\s*territor(?:y|ies)\s*:\s*([^\r\n]+)", re.IGNORECASE),
    re.compile(r"territor(?:y|ies)\s*requested\s*:\s*([^\r\n]+)", re.IGNORECASE),
    re.compile(r"territor(?:y|ies)\s*:\s*([^\r\n]+)", re.IGNORECASE),
    re.compile(r"location\s*:\s*([^\r\n]+)", re.IGNORECASE),
)
ANYWHERE_TERR_FALLBACK = re.compile(
    r"territor(?:y|ies)[^A-Za-z]{0,20}([A-Za-z .'\-]+,\s*[A-Za-z]{2})", re.IGNORECASE | re.DOTALL
)

def territory_from_subject(subject: str) -> str:
    s = _norm(subject)
    for pat in TERR_PATTERNS:
        m = pat.search(s)
        if m: return m.group(1).title()
    return ""

def _clean_territory(raw: str) -> str:
    s = _norm(raw or "")
    s = re.split(r"[|/;\n\r]", s)[0].strip()
    s = re.sub(r"\s+", " ", s)
    return s.title()

def territory_from_any(subject: str, body_preview: str, full_body_text: str | None = None) -> str:
    terr = territory_from_subject(subject)
    if terr: return terr
    for pat in BODY_TERR_PATTERNS:
        m = pat.search(body_preview or "")
        if m: return _clean_territory(m.group(1))
    m = ANYWHERE_TERR_FALLBACK.search(body_preview or "")
    if m: return _clean_territory(m.group(1))
    if full_body_text:
        for pat in BODY_TERR_PATTERNS:
            m = pat.search(full_body_text)
            if m: return _clean_territory(m.group(1))
        m = ANYWHERE_TERR_FALLBACK.search(full_body_text)
        if m: return _clean_territory(m.group(1))
        m = re.search(r"\b([A-Za-z .'\-]+,\s*[A-Za-z]{2})\b", full_body_text)
        if m: return _clean_territory(m.group(1))
    return ""

def pretty_label_from_domain(email_or_domain: str) -> str:
    dom = (email_or_domain or "").split("@")[-1].lower()
    if not dom or "." not in dom: return "Unknown"
    core = dom.split(".")[0]
    chunks = re.split(r"[\W_]+", core)
    if not chunks or chunks == [""]: return core.title()
    return " ".join([c.upper() if len(c) <= 3 else c.title() for c in chunks if c])

# ================================ Excel Helpers ===============================
def ensure_out_dir(): os.makedirs(OUT_DIR, exist_ok=True)

def unique_week_filepath(end_label: str) -> str:
    base = os.path.join(OUT_DIR, f"Territory_Checks_{end_label}.xlsx")
    if not os.path.exists(base): return base
    i = 1
    while True:
        cand = os.path.join(OUT_DIR, f"Territory_Checks_{end_label} ({i}).xlsx")
        if not os.path.exists(cand): return cand
        i += 1

def try_open_workbook(path: str) -> Optional[Workbook]:
    if not os.path.exists(path):
        wb = Workbook(); wb.active.title = "Index"; return wb
    try: return load_workbook(path)
    except PermissionError: return None

# ================================== Main ======================================
def run() -> None:
    # ---- Env validation
    missing = [k for k, v in {
        "GRAPH_TENANT_ID": TENANT_ID, "GRAPH_CLIENT_ID": CLIENT_ID,
        "GRAPH_CLIENT_SECRET": CLIENT_SECRET, "MASTER_MAILBOX": MASTER_MAILBOX, "OUT_DIR": OUT_DIR,
    }.items() if not v]
    if missing:
        raise SystemExit("Missing environment variables: " + ", ".join(missing))

    ensure_out_dir()
    start_dt, end_dt, end_excl, week_label, file_label = last_completed_week_utc()
    log.info("Processing last completed week: %s → %s (%s)", start_dt, end_dt, week_label)

    client = GraphClient(TENANT_ID, CLIENT_ID, CLIENT_SECRET, MASTER_MAILBOX)
    client.authenticate()

    start_iso = start_dt.strftime("%Y-%m-%dT%H:%M:%SZ")
    end_iso   = end_excl.strftime("%Y-%m-%dT%H:%M:%SZ")

    details_rows: List[dict] = []
    audit_rows:   List[dict] = []

    # ---- Scan ALL folders
    for f in client.list_all_folders():
        fid = f["id"]; fname = f.get("displayName", "")
        log.info("Scanning folder: %s", fname)
        for m in client.fetch_messages_in_folder(fid, start_iso, end_iso):
            subj = (m.get("subject") or "").strip()
            if SKIP_REPLIES and is_reply(subj):
                audit_rows.append({"Folder": fname, "Reason": "Dropped: Reply", "Subject": subj})
                continue
            if not COUNT_FORWARDS and is_forward(subj):
                audit_rows.append({"Folder": fname, "Reason": "Dropped: Forward (disabled)", "Subject": subj})
                continue

            # ----- Brand: To/Cc/Bcc → headers → Unmapped
            b_to_cc_bcc, source1 = brand_from_recipients(m.get("toRecipients"), m.get("ccRecipients"), m.get("bccRecipients"))
            b_hdrs, source2 = (None, "")
            if not b_to_cc_bcc:
                b_hdrs, source2 = brand_from_headers(m.get("internetMessageHeaders"))
            brand = b_to_cc_bcc or b_hdrs or "Unmapped"
            brand_source = source1 or source2 or "Unmapped"

            # ----- Sender / preview
            from_addr    = ((m.get("from") or {}).get("emailAddress") or {})
            sender_email = (from_addr.get("address") or "").strip()
            sender_name  = (from_addr.get("name") or "").strip() or sender_email
            body_preview = (m.get("bodyPreview") or "").strip()

            # ----- Broker: subject → forwarded header → domain
            broker_subj = broker_from_subject(subj)
            broker_fw   = None if broker_subj else forwarded_broker_from_bodypreview(body_preview)
            broker_dom  = broker_from_sender(m.get("from"))
            chosen_broker = broker_subj or broker_fw or broker_dom

            # If still Others, produce friendly label in details
            detail_broker_label = chosen_broker
            others_label = None
            if chosen_broker == "Others":
                pretty = pretty_label_from_domain(sender_email)
                detail_broker_label = f"Others | {pretty}"
                others_label = pretty

            # ----- Territory anywhere: subject → bodyPreview → full body (lazy)
            terr = territory_from_any(subj, body_preview, None)
            full_text = None
            if not terr:
                try:
                    body_html = client.fetch_full_body(m["id"])
                    full_text = _html_to_text(body_html)
                    terr = territory_from_any(subj, body_preview, full_text)
                except Exception as e:
                    log.debug("Full body fetch failed for %s: %s", m.get("id"), e)

            recv    = m.get("receivedDateTime", "")
            recv_dt = recv[:10] if recv else ""

            details_rows.append({
                "Brand"              : brand,
                "Broker Brand"       : detail_broker_label,
                "Summary Bucket"     : chosen_broker,
                "Others Name"        : others_label or "",
                "Broker Name"        : sender_name,
                "Territory"          : terr,
                "Received (UTC Date)": recv_dt,
                "Subject"            : subj,
                "From"               : sender_email,
                "Folder"             : fname,
                "Brand Source"       : brand_source,
                "ConversationId"     : m.get("conversationId", ""),
                "InternetMessageId"  : m.get("internetMessageId", ""),
            })

            audit_rows.append({
                "Folder"               : fname,
                "Brand"                : brand,
                "Brand Source"         : brand_source,
                "IsForward"            : is_forward(subj),
                "IsReply"              : is_reply(subj),
                "Subject"              : subj,
                "From"                 : sender_email,
                "To"                   : ";".join([(x.get("emailAddress") or {}).get("address","") for x in (m.get("toRecipients") or [])]),
                "CC"                   : ";".join([(x.get("emailAddress") or {}).get("address","") for x in (m.get("ccRecipients") or [])]),
                "BCC"                  : ";".join([(x.get("emailAddress") or {}).get("address","") for x in (m.get("bccRecipients") or [])]),
                "BodyPreview"          : body_preview,
                "ReceivedUTC"          : recv,
                "FetchedFullBody"      : bool(full_text),
                "Chosen Broker (bucket)": chosen_broker,
            })

    # ---- Build DataFrame & light duplicate guard
    df = pd.DataFrame(details_rows)
    if df.empty:
        log.info("No messages in the window; writing an empty workbook with all brand sheets.")
        # still emit an empty workbook with headers for every brand + Unmapped + Audit
        _write_workbook(pd.DataFrame(columns=[
            "Brand","Broker Brand","Summary Bucket","Others Name","Broker Name","Territory",
            "Received (UTC Date)","Subject","From","Folder","Brand Source",
            "ConversationId","InternetMessageId"
        ]), audit_rows, week_label, file_label)
        return

    if "InternetMessageId" in df.columns:
        df = df.sort_values(by=["InternetMessageId", "Received (UTC Date)"])
        df = df.drop_duplicates(subset=["InternetMessageId"], keep="first")

    # ---- De-duplicate by Brand+Broker+Territory (when territory known)
    def _canon_terr(x: str) -> str:
        s = _norm(x or "")
        s = re.sub(r"[^a-z0-9 ,.'\-]", " ", s)
        s = re.sub(r"\s+", " ", s).strip()
        return s
    df["__TerrKey__"] = df["Territory"].map(_canon_terr)
    with_terr    = df[df["__TerrKey__"] != ""]
    without_terr = df[df["__TerrKey__"] == ""]
    deduped = with_terr.drop_duplicates(
        subset=["Brand", "Summary Bucket", "Broker Brand", "__TerrKey__"],
        keep="first"
    )
    df = pd.concat([deduped, without_terr], ignore_index=True)

    # ---- Sort & sequence
    df["Received (UTC Date)"] = pd.to_datetime(df["Received (UTC Date)"], errors="coerce")
    df.sort_values(by=["Brand", "Summary Bucket", "Received (UTC Date)", "Subject"], inplace=True)
    df["Seq"] = df.groupby(["Brand", "Broker Brand"]).cumcount() + 1

    _write_workbook(df, audit_rows, week_label, file_label)

# ------------------------------- Writer ---------------------------------------
def _write_workbook(df: pd.DataFrame, audit_rows: List[dict], week_label: str, file_label: str) -> None:
    seen_brokers = [b for b in df.get("Summary Bucket", pd.Series([], dtype=str)).unique().tolist()
                    if b not in BROKER_ORDER and b not in (None, "Unmapped")]
    broker_order_full = BROKER_ORDER + sorted([b for b in seen_brokers if isinstance(b, str)])

    target_path = unique_week_filepath(file_label)
    wb = try_open_workbook(target_path)
    if wb is None:
        i = 1
        while wb is None:
            alt = os.path.join(OUT_DIR, f"Territory_Checks_{file_label} ({i}).xlsx")
            wb = try_open_workbook(alt)
            if wb:
                target_path = alt
                break
            i += 1

    # ---- Per-brand data sheets
    if not df.empty:
        for brand, sub in df.groupby("Brand"):
            sheet = brand
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet)
            ws = wb[sheet]
            ws.delete_rows(1, ws.max_row)

            cols_detail = ["Broker Brand", "Broker Name", "Territory", "Received (UTC Date)", "Seq"]
            if SHOW_SUBJECT_IN_DETAILS:
                cols_detail.append("Subject")

            sub_out = sub[cols_detail + ["Summary Bucket"]].copy()
            sub_out["__order__"] = sub_out["Summary Bucket"].apply(
                lambda b: broker_order_full.index(b) if b in broker_order_full else 999
            )
            sub_out.sort_values(by=["__order__", "Received (UTC Date)", "Seq"], inplace=True)
            sub_out.drop(columns=["__order__", "Summary Bucket"], inplace=True)

            ws.append(cols_detail)
            for _, r in sub_out.iterrows():
                ws.append([r[c] for c in cols_detail])

            # Summary (start at G)
            col_base = 7
            summary = sub.groupby("Summary Bucket").size().rename("Count").reset_index()
            summary = summary.set_index("Summary Bucket").reindex(broker_order_full, fill_value=0).reset_index()
            total_row = pd.DataFrame([{"Summary Bucket": "Total", "Count": int(summary["Count"].sum())}])
            summary = pd.concat([summary, total_row], ignore_index=True)

            ws.cell(row=1, column=col_base,   value="Broker Brand")
            ws.cell(row=1, column=col_base+1, value=week_label)
            for i, row in summary.iterrows():
                ws.cell(row=i+2, column=col_base,   value=row["Summary Bucket"])
                ws.cell(row=i+2, column=col_base+1, value=int(row["Count"]))

            # Others breakdown (below)
            others_sub = sub[sub["Summary Bucket"] == "Others"]
            if not others_sub.empty:
                breakdown = (others_sub
                             .assign(Detail=others_sub["Broker Brand"].str.replace(r"^Others \|\s*", "", regex=True))
                             .groupby("Detail").size().rename("Count").reset_index()
                             .sort_values(by=["Count","Detail"], ascending=[False, True]))
                start_r = len(summary) + 4
                ws.cell(row=start_r-1, column=col_base,   value="Others breakdown")
                ws.cell(row=start_r,   column=col_base,   value="Detail")
                ws.cell(row=start_r,   column=col_base+1, value="Count")
                for j, row in breakdown.iterrows():
                    ws.cell(row=start_r + j + 1, column=col_base,   value=row["Detail"])
                    ws.cell(row=start_r + j + 1, column=col_base+1, value=int(row["Count"]))

    # ---- Ensure every brand has a sheet, even with 0 rows
    cols_detail = ["Broker Brand", "Broker Name", "Territory", "Received (UTC Date)", "Seq"]
    if SHOW_SUBJECT_IN_DETAILS:
        cols_detail.append("Subject")

    for brand_name in BRAND_ADDRESSES.keys():
        if brand_name not in wb.sheetnames:
            wb.create_sheet(brand_name)
        ws = wb[brand_name]
        # If sheet empty, lay down headers + zero summary
        if ws.max_row <= 1 and ws.max_column <= 1:
            ws.delete_rows(1, ws.max_row)
            ws.append(cols_detail)
            col_base = 7
            ws.cell(row=1, column=col_base,   value="Broker Brand")
            ws.cell(row=1, column=col_base+1, value=week_label)
            # zero summary for all brokers + Total
            for i, b in enumerate(broker_order_full, start=2):
                ws.cell(row=i, column=col_base,   value=b)
                ws.cell(row=i, column=col_base+1, value=0)
            ws.cell(row=len(broker_order_full)+2, column=col_base,   value="Total")
            ws.cell(row=len(broker_order_full)+2, column=col_base+1, value=0)

    # ---- Unmapped
    if "Unmapped" not in wb.sheetnames:
        wb.create_sheet("Unmapped")
    ws_u = wb["Unmapped"]
    ws_u.delete_rows(1, ws_u.max_row)
    cols_u = ["Brand", "Brand Source", "Broker Brand", "Broker Name", "Territory",
              "Received (UTC Date)", "Subject", "From", "Folder"]
    ws_u.append(cols_u)
    if not df.empty and "Brand" in df.columns:
        for _, r in df[df["Brand"] == "Unmapped"][cols_u].iterrows():
            ws_u.append([r.get(c, "") for c in cols_u])

    # ---- Audit
    if "Audit" not in wb.sheetnames:
        wb.create_sheet("Audit")
    ws_a = wb["Audit"]
    ws_a.delete_rows(1, ws_a.max_row)
    cols_a = ["Folder","Brand","Brand Source","IsForward","IsReply","Subject","From","To","CC","BCC",
              "BodyPreview","ReceivedUTC","FetchedFullBody","Chosen Broker (bucket)"]
    ws_a.append(cols_a)
    for row in audit_rows:
        ws_a.append([row.get(c, "") for c in cols_a])

    # ---- Remove placeholder
    if "Index" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Index"]

    # ---- Save (lock-safe)
    try:
        wb.save(target_path)
        log.info("Saved weekly workbook: %s", target_path)
    except PermissionError:
        j = 1
        while True:
            alt2 = os.path.join(OUT_DIR, f"Territory_Checks_{file_label} (save {j}).xlsx")
            try:
                wb.save(alt2)
                log.info("Target locked; saved as: %s", alt2)
                break
            except PermissionError:
                j += 1

# ================================= Entrypoint =================================
if __name__ == "__main__":
    try:
        run()
    except SystemExit as e:
        log.error(str(e)); raise
    except Exception:
        log.exception("Unhandled error"); raise
