import io
import os
import re
from datetime import datetime
from pathlib import Path

from google.cloud import bigquery, storage
import openpyxl as xl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.utils.units import pixels_to_EMU


# ----------------------------
# CONFIG
# ----------------------------
BQ_PROJECT = os.getenv("BQ_PROJECT", "sfs-data-lake")
BQ_DATASET = os.getenv("BQ_DATASET", "silver")
BQ_TABLE = os.getenv("BQ_TABLE", "sfs_internal_announcements")
BALANCE_SUMMARY_TYPES = {
    "BALANCE",
    "DEPOSIT + BALANCE",
}

TARGET_YEAR = int(os.getenv("TARGET_YEAR", "2026"))
TEMPLATE_SHEET = os.getenv("TEMPLATE_SHEET", f"{TARGET_YEAR} BALANCES REPORT")

GCS_BUCKET = os.getenv("GCS_BUCKET", "sfs-raw-us")
GCS_TEMPLATE_BLOB = os.getenv(
    "GCS_TEMPLATE_BLOB",
    "sfs_strategic_franchising/Balances_VS_Budget/templates/2026_Balances_vs_Budget_template.xlsx",
)
GCS_OUTPUT_PREFIX = os.getenv(
    "GCS_OUTPUT_PREFIX",
    "sfs_strategic_franchising/Balances_VS_Budget/outputs",
)

LOGO_DIR = os.getenv(
    "LOGO_DIR",
    os.path.join(os.path.dirname(__file__), "logos"),
)
LOGO_ANCHOR_COL = os.getenv("LOGO_ANCHOR_COL", "B")
LOGO_Y_OFFSET_PX = int(os.getenv("LOGO_Y_OFFSET_PX", "-2"))
LOGO_SHEETS = [
    s.strip()
    for s in os.getenv(
        "LOGO_SHEETS",
        "2026 BALANCES REPORT,2025 BALANCES REPORT",
    ).split(",")
    if s.strip()
]

# 2026 template row mapping (ACTUAL rows only)
ROW_MAP = {
    "ORGANIC": {
        "Caring Transitions": 32,
        "Fresh Coat": 37,
        "Growth Coach": 42,
        "Pet Wants": 47,
        "TruBlue": 52,
    },
    "BROKER": {
        "Caring Transitions": 62,
        "Fresh Coat": 67,
        "Growth Coach": 72,
        "Pet Wants": 77,
        "TruBlue": 82,
    },
}

# Expected logo filenames in LOGO_DIR.
LOGO_FILES = {
    "Caring Transitions": ["CT.png"],
    "Fresh Coat": ["FC.png"],
    "Growth Coach": ["GC.jpeg", "GC.jpg"],
    "Pet Wants": ["PW.jpeg", "PW.jpg"],
    "TruBlue": ["TB.jpeg", "TB.jpg"],
    "Strategic Franchising": ["SFS.png"],
}

# Max rendered size (pixels) per brand logo to keep layout consistent in Excel.
LOGO_MAX_SIZE = {
    "Caring Transitions": (170, 58),
    "Fresh Coat": (145, 58),       # smaller
    "Growth Coach": (330, 58),     # wider (default tuned)
    "Pet Wants": (330, 58),        # wider (default tuned)
    "TruBlue": (185, 60),
    "Strategic Franchising": (330, 60),
}

# Horizontal shift (pixels) to visually center logos in brand column.
LOGO_X_OFFSET = {
    "default": 30,
    "Caring Transitions": 30,
    "Fresh Coat": 32,
    "Growth Coach": 34,
    "Pet Wants": 34,
    "TruBlue": 30,
    "Strategic Franchising": 32,
}

# Row anchors for each section in template.
LOGO_ROW_GROUPS = [
    {
        "Caring Transitions": 3,
        "Fresh Coat": 8,
        "Growth Coach": 13,
        "Pet Wants": 18,
        "TruBlue": 23,
        "Strategic Franchising": 28,
    },
    {
        "Caring Transitions": 32,
        "Fresh Coat": 37,
        "Growth Coach": 42,
        "Pet Wants": 47,
        "TruBlue": 52,
        "Strategic Franchising": 57,
    },
    {
        "Caring Transitions": 62,
        "Fresh Coat": 67,
        "Growth Coach": 72,
        "Pet Wants": 77,
        "TruBlue": 82,
        "Strategic Franchising": 87,
    },
]


def load_template_from_gcs(bucket: str, blob_name: str) -> xl.Workbook:
    client = storage.Client()
    b = client.bucket(bucket)
    blob = b.blob(blob_name)
    data = blob.download_as_bytes()
    return xl.load_workbook(io.BytesIO(data))


def upload_to_gcs(bucket: str, blob_name: str, local_path: str) -> None:
    client = storage.Client()
    b = client.bucket(bucket)
    blob = b.blob(blob_name)
    blob.upload_from_filename(local_path)


def month_col(month_num: int) -> int:
    # D..O => Jan..Dec
    return 3 + month_num


def fetch_actual_counts() -> dict[tuple[str, str, int], int]:
    allowed_types_sql = ", ".join(f"'{t}'" for t in sorted(BALANCE_SUMMARY_TYPES))
    sql = f"""
    SELECT
      EXTRACT(MONTH FROM closed_sale_date) AS month_num,
      CASE
        WHEN UPPER(TRIM(brand)) IN ('CT', 'CARING TRANSITIONS') THEN 'Caring Transitions'
        WHEN UPPER(TRIM(brand)) IN ('FC', 'FRESH COAT', 'FRESH COAT PAINTERS') THEN 'Fresh Coat'
        WHEN UPPER(TRIM(brand)) IN ('GC', 'GROWTH COACH', 'THE GROWTH COACH') THEN 'Growth Coach'
        WHEN UPPER(TRIM(brand)) IN ('PW', 'PET WANTS') THEN 'Pet Wants'
        WHEN UPPER(TRIM(brand)) IN ('TB', 'TRUBLUE', 'TRU BLUE', 'TRU BLUE ALLY') THEN 'TruBlue'
        ELSE TRIM(brand)
      END AS brand_norm,
      CASE
        WHEN LOWER(COALESCE(lead_source, '')) LIKE 'broker%' THEN 'BROKER'
        ELSE 'ORGANIC'
      END AS channel,
      COUNT(
        DISTINCT COALESCE(
          NULLIF(franchisee_id, ''),
          CONCAT('NID|', franchisee_name, '|', CAST(closed_sale_date AS STRING))
        )
      ) AS actual_count
    FROM `{BQ_PROJECT}.{BQ_DATASET}.{BQ_TABLE}`
    WHERE closed_sale_date IS NOT NULL
      AND balance_deposit_date IS NOT NULL
      AND UPPER(TRIM(COALESCE(announcement_type, ''))) IN ({allowed_types_sql})
      AND EXTRACT(YEAR FROM closed_sale_date) = @target_year
    GROUP BY month_num, brand_norm, channel
    """

    client = bigquery.Client(project=BQ_PROJECT)
    job_config = bigquery.QueryJobConfig(
        query_parameters=[
            bigquery.ScalarQueryParameter("target_year", "INT64", TARGET_YEAR),
        ]
    )
    rows = client.query(sql, job_config=job_config).result()

    out: dict[tuple[str, str, int], int] = {}
    for r in rows:
        key = (r["brand_norm"], r["channel"], int(r["month_num"]))
        out[key] = int(r["actual_count"])
    return out


def write_actuals(ws: xl.worksheet.worksheet.Worksheet, counts: dict[tuple[str, str, int], int]) -> None:
    # Fill D..O for brand ACTUAL rows in ORGANIC/BROKER sections.
    for channel, brand_rows in ROW_MAP.items():
        for brand, row in brand_rows.items():
            for m in range(1, 13):
                ws.cell(row=row, column=month_col(m), value=counts.get((brand, channel, m), 0))


def _logo_index() -> dict[str, Path]:
    logo_dir = Path(LOGO_DIR)
    if not logo_dir.exists():
        return {}
    return {
        p.name.lower(): p
        for p in logo_dir.iterdir()
        if p.is_file()
    }


def _find_logo(brand: str, index: dict[str, Path]) -> Path | None:
    for name in LOGO_FILES.get(brand, []):
        p = index.get(name.lower())
        if p:
            return p
    return None


def _brand_token(brand: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "_", brand.upper()).strip("_")


def _apply_logo_size(img: XLImage, brand: str) -> None:
    max_w, max_h = LOGO_MAX_SIZE.get(brand, (170, 60))
    token = _brand_token(brand)
    max_w = int(os.getenv(f"LOGO_{token}_W", str(max_w)))
    max_h = int(os.getenv(f"LOGO_{token}_H", str(max_h)))
    ow, oh = float(img.width), float(img.height)
    if ow <= 0 or oh <= 0:
        return
    scale = min(max_w / ow, max_h / oh)
    img.width = max(1, int(ow * scale))
    img.height = max(1, int(oh * scale))


def _set_logo_anchor(img: XLImage, row: int, brand: str) -> None:
    col_idx = column_index_from_string(LOGO_ANCHOR_COL) - 1
    x_px = LOGO_X_OFFSET.get(brand, LOGO_X_OFFSET["default"])
    token = _brand_token(brand)
    x_px = int(os.getenv(f"LOGO_{token}_X", str(x_px)))
    marker = AnchorMarker(
        col=col_idx,
        row=row - 1,
        colOff=pixels_to_EMU(x_px),
        rowOff=pixels_to_EMU(LOGO_Y_OFFSET_PX),
    )
    ext = XDRPositiveSize2D(
        cx=pixels_to_EMU(int(img.width)),
        cy=pixels_to_EMU(int(img.height)),
    )
    img.anchor = OneCellAnchor(_from=marker, ext=ext)


def _load_logo_image(logo_path: Path) -> XLImage:
    # Keep file-backed image objects so openpyxl can serialize reliably on save.
    return XLImage(str(logo_path))


def restore_logos(wb: xl.Workbook) -> None:
    index = _logo_index()
    if not index:
        print(f"[WARN] LOGO_DIR not found or empty: {LOGO_DIR}")
        return

    for sheet_name in LOGO_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ws._images = []

        added = 0
        for group in LOGO_ROW_GROUPS:
            for brand, row in group.items():
                logo_path = _find_logo(brand, index)
                if not logo_path:
                    print(f"[WARN] Missing logo for {brand} in {LOGO_DIR}")
                    continue
                img = _load_logo_image(logo_path)
                _apply_logo_size(img, brand)
                _set_logo_anchor(img, row, brand)
                ws.add_image(img)
                added += 1
        print(f"[INFO] {sheet_name}: reinserted logos={added}")


def apply_month_visibility(ws: xl.worksheet.worksheet.Worksheet, counts: dict[tuple[str, str, int], int]) -> None:
    populated_months = {m for (_, _, m), _v in counts.items() if 1 <= m <= 12}
    latest = max(populated_months) if populated_months else 1
    for m in range(1, 13):
        col_letter = get_column_letter(month_col(m))
        ws.column_dimensions[col_letter].hidden = m > latest
    print(f"[INFO] {ws.title}: showing months 1..{latest}, hiding {latest + 1}..12")


def main() -> None:
    print("[INFO] Loading template from GCS...")
    wb = load_template_from_gcs(GCS_BUCKET, GCS_TEMPLATE_BLOB)
    if TEMPLATE_SHEET not in wb.sheetnames:
        raise ValueError(f"Template sheet not found: {TEMPLATE_SHEET}")
    ws = wb[TEMPLATE_SHEET]

    print("[INFO] Fetching ACTUAL counts from BigQuery...")
    counts = fetch_actual_counts()
    print(f"[INFO] Rows returned for year {TARGET_YEAR}: {len(counts)}")

    print("[INFO] Writing ACTUAL cells (ORGANIC/BROKER)...")
    write_actuals(ws, counts)
    apply_month_visibility(ws, counts)
    restore_logos(wb)

    out_name = f"balances_vs_budget_{TARGET_YEAR}_{datetime.utcnow():%Y-%m-%d}.xlsx"
    local_path = f"/tmp/{out_name}"
    wb.save(local_path)

    gcs_out = f"{GCS_OUTPUT_PREFIX}/{datetime.utcnow():%Y/%m}/{out_name}"
    upload_to_gcs(GCS_BUCKET, gcs_out, local_path)
    print(f"[INFO] Uploaded: gs://{GCS_BUCKET}/{gcs_out}")


if __name__ == "__main__":
    main()
