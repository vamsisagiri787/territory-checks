import os
import re
import io
import calendar
from datetime import datetime
import pandas as pd
from google.cloud import bigquery, storage
import openpyxl as xl
from openpyxl.cell.cell import MergedCell
from copy import copy

# ----------------------------
# CONFIG
# ----------------------------
BQ_PROJECT = os.getenv("BQ_PROJECT", "sfs-data-lake")
BQ_DATASET = os.getenv("BQ_DATASET", "silver")
BQ_TABLE = os.getenv("BQ_TABLE", "sfs_internal_announcements")

GCS_BUCKET = os.getenv("GCS_BUCKET", "sfs-raw-us")
GCS_TEMPLATE_BLOB = os.getenv(
    "GCS_TEMPLATE_BLOB",
    "sfs_strategic_franchising/Balances and Deposits/templates/Gold_Template_Internal_Announcements.xlsx"
)
# ===================== SFS_009 =====================
# Reason: Keep workbook uploads under the same weekly/year/month folder pattern
# used by territory checks so bucket navigation stays consistent across reports.
# ====================================================
GCS_OUTPUT_PREFIX = os.getenv(
    "GCS_OUTPUT_PREFIX",
    "sfs_strategic_franchising/Balances and Deposits/outputs/weekly"
)

BRAND_SHEET = {
    "Caring Transitions": "CT",
    "TruBlue": "TB",
    "Fresh Coat": "FC",
    "Growth Coach": "GC",
    "Pet Wants": "PW",
    "CT": "CT",
    "TB": "TB",
    "FC": "FC",
    "GC": "GC",
    "PW": "PW",
}

# ----------------------------
# QUERY
# ----------------------------
SQL = f"""
SELECT
  brand,
  director_of_franchising,
  franchisee_name,
  franchisee_id,
  state_code,
  lead_source,
  announcement_type,
  balance_deposit_date,
  closed_sale_date,
  amount_usd,
  received_datetime,
  FORMAT_DATE('%Y-%m', COALESCE(closed_sale_date, balance_deposit_date)) AS signing_month
FROM `{BQ_PROJECT}.{BQ_DATASET}.{BQ_TABLE}`
WHERE COALESCE(closed_sale_date, balance_deposit_date) IS NOT NULL
"""

# ----------------------------
# HELPERS
# ----------------------------
def load_template_from_gcs(bucket, blob_name):
    client = storage.Client()
    b = client.bucket(bucket)
    blob = b.blob(blob_name)
    data = blob.download_as_bytes()
    return xl.load_workbook(io.BytesIO(data))

def upload_to_gcs(bucket, blob_name, local_path):
    client = storage.Client()
    b = client.bucket(bucket)
    blob = b.blob(blob_name)
    blob.upload_from_filename(local_path)

def clear_block(ws, start_row):
    # Remove all rows below header to avoid stale gaps
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)
    # Remove merged ranges below header rows (if any remain)
    to_unmerge = [m for m in ws.merged_cells.ranges if m.min_row >= start_row]
    for m in to_unmerge:
        ws.unmerge_cells(str(m))

def _copy_style(src, dest):
    # copy() to avoid StyleProxy unhashable errors
    dest.font = copy(src.font)
    dest.fill = copy(src.fill)
    dest.border = copy(src.border)
    dest.alignment = copy(src.alignment)
    dest.number_format = src.number_format

def write_rows(ws, start_row, rows):
    r = start_row
    last_written = None
    for row in rows:
        ann_type = row.get("announcement_type")
        if not ann_type:
            ann_type = row.get("lead_source") or "OTHER"
        # Normalize amount to numeric for Excel sums
        amt = row.get("amount_usd")
        if isinstance(amt, str):
            amt = amt.replace(",", "").replace("$", "").strip()
            try:
                amt = float(amt) if amt else None
            except Exception:
                amt = None
        elif amt is not None:
            try:
                amt = float(amt)
            except Exception:
                pass
        ws.cell(r, 1, row.get("director_of_franchising"))
        ws.cell(r, 2, row.get("franchisee_name"))
        ws.cell(r, 3, row.get("franchisee_id"))
        ws.cell(r, 4, row.get("state_code"))
        ws.cell(r, 5, row.get("lead_source"))
        ws.cell(r, 6, ann_type)
        date_cell = ws.cell(r, 7, row.get("balance_deposit_date"))
        date_cell.number_format = "M/D/YYYY"
        close_cell = ws.cell(r, 8, row.get("closed_sale_date"))
        close_cell.number_format = "M/D/YYYY"
        amt_cell = ws.cell(r, 9, amt)
        amt_cell.number_format = '"$"#,##0.00'
        last_written = r
        r += 1
    return last_written

def is_balance(atype):
    if not atype:
        return False
    a = str(atype).upper()
    # Partial balance should go below totals (other section)
    if "PARTIAL BALANCE" in a:
        return False
    # If it includes BALANCE (even with DEPOSIT), keep in balances section
    return "BALANCE" in a

def _is_action_type(v: str) -> bool:
    s = (v or "").strip().upper()
    if not s:
        return False
    keys = [
        "TRANSFER COMPLETE",
        "TRAINING APPROVED/CLOSED DEAL",
        "DEAL CLOSED",
        "ADDITIONAL FRANCHISE PURCHASE COMPLETE",
        "ROFR AGREEMENT SIGNED",
    ]
    return any(k in s for k in keys)

# ----------------------------
# MAIN
# ----------------------------
def main():
    # Query silver
    bq = bigquery.Client(project=BQ_PROJECT)
    df = bq.query(SQL).to_dataframe(create_bqstorage_client=False)
    if not df.empty:
        # Normalize for robust dedupe/filtering.
        df["brand"] = df["brand"].fillna("").astype(str).str.strip()
        df["franchisee_id"] = df["franchisee_id"].fillna("").astype(str).str.strip()
        df["franchisee_name"] = df["franchisee_name"].fillna("").astype(str).str.strip()
        df["announcement_type"] = df["announcement_type"].fillna("").astype(str).str.strip()
        df["received_datetime"] = pd.to_datetime(df["received_datetime"], errors="coerce")

        # ===================== SFS_004 =====================
        # Reason: silver can still contain closed-date-only residue rows from
        # ACTION emails after the canonical financial row already exists. Some
        # of those residual rows now carry announcement_type=OTHER after merge,
        # so suppress them in the workbook layer using the business shape of the
        # row rather than announcement_type alone.
        # ==================================================
        has_core = (
            df["balance_deposit_date"].notna()
            | df["amount_usd"].notna()
            | ~df["announcement_type"].apply(_is_action_type)
        )
        key_id = (df["brand"] + "|" + df["franchisee_id"]).str.strip("|")
        key_name = (df["brand"] + "|" + df["franchisee_name"]).str.strip("|")
        core_ids = set(key_id[(df["franchisee_id"] != "") & has_core].tolist())
        core_names = set(key_name[(df["franchisee_id"] == "") & (df["franchisee_name"] != "") & has_core].tolist())

        is_action_only = df["announcement_type"].apply(_is_action_type) & df["balance_deposit_date"].isna() & df["amount_usd"].isna()
        is_residual_other = (
            df["announcement_type"].fillna("").astype(str).str.strip().str.upper().eq("OTHER")
            & df["balance_deposit_date"].isna()
            & df["amount_usd"].isna()
            & df["closed_sale_date"].notna()
        )
        drop_mask = (
            (is_action_only | is_residual_other)
            & (
                ((df["franchisee_id"] != "") & key_id.isin(core_ids))
                | ((df["franchisee_id"] == "") & (df["franchisee_name"] != "") & key_name.isin(core_names))
            )
        )
        df = df[~drop_mask].copy()

        # Keep latest row per business event, not just per franchise/month/type.
        # Reason: the same franchise can legitimately have multiple rows in the same
        # month/type (for example Rebecca + Jeff partial-balance rows with different
        # dates/amounts). We still want to collapse exact duplicates, so include the
        # financial/date fields in the dedupe key.
        dedupe_key = (
            df["brand"] + "|" + df["franchisee_id"] + "|" + df["franchisee_name"] + "|"
            + df["signing_month"].fillna("") + "|" + df["announcement_type"] + "|"
            + df["balance_deposit_date"].astype(str).fillna("") + "|"
            + df["closed_sale_date"].astype(str).fillna("") + "|"
            + df["amount_usd"].astype(str).fillna("")
        )
        df["_dedupe_key"] = dedupe_key
        df = (
            df.sort_values(by=["_dedupe_key", "received_datetime"], na_position="last")
            .groupby("_dedupe_key", as_index=False)
            .tail(1)
            .drop(columns=["_dedupe_key"])
        )

    # Normalize brand codes to match template sheets
    df["brand"] = df["brand"].replace({
        "CT": "Caring Transitions",
        "TB": "TruBlue",
        "FC": "Fresh Coat",
        "GC": "Growth Coach",
        "PW": "Pet Wants",
    })

    # Load template
    wb = load_template_from_gcs(GCS_BUCKET, GCS_TEMPLATE_BLOB)
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    # Fill per brand
    for brand, sheet in BRAND_SHEET.items():
        if sheet not in wb.sheetnames:
            continue

        ws = wb[sheet]
        df_b = df[df["brand"] == brand].copy()
        if df_b.empty:
            continue

        # clear existing rows below header
        clear_block(ws, start_row=3)

        current_row = 3

        # group by month
        for month, df_m in df_b.groupby("signing_month"):
            dt = datetime.strptime(month, "%Y-%m")
            month_name = calendar.month_name[dt.month]
            yy = f"{dt.year % 100:02d}"

            # Month header row (merged across 9 columns) styled like brand header
            month_header = f"{month_name} {yy}"
            header_row = current_row
            ws.merge_cells(start_row=header_row, start_column=1, end_row=header_row, end_column=9)
            cell = ws.cell(header_row, 1, month_header)
            _copy_style(ws.cell(1, 1), cell)

            # Split balances vs other
            balances = df_m[df_m["announcement_type"].apply(is_balance)]
            other = df_m[~df_m["announcement_type"].apply(is_balance)]

            # Layout:
            # balances rows -> 1 blank row -> total row -> 1 blank row -> other rows
            balances_start = header_row + 1

            last_balance_row = write_rows(
                ws, balances_start, balances.to_dict("records")
            )
            print(
                f"[INFO] {brand} {month_name}: balances={len(balances)} "
                f"rows_start={balances_start} last_row={last_balance_row}"
            )

            # Total row only when balances exist.
            if last_balance_row is not None:
                total_row = last_balance_row + 2
                ws.cell(total_row, 1, f"Total {month_name} Balances Collected")
                total_cell = ws.cell(total_row, 9)
                total_cell.value = f"=SUM(I{balances_start}:I{last_balance_row})"
                try:
                    total_cell.data_type = "f"
                except Exception:
                    pass
                total_cell.number_format = '"$"#,##0.00'
                print(
                    f"[INFO] {brand} {month_name}: total_row={total_row} "
                    f"formula={total_cell.value}"
                )
                other_start = total_row + 2  # 1 blank row after total
            else:
                print(f"[INFO] {brand} {month_name}: NO balances, skip total row")
                other_start = balances_start

            write_rows(ws, other_start, other.to_dict("records"))

            # Advance current_row to the next free row after other rows
            if other.empty:
                current_row = other_start
            else:
                current_row = other_start + len(other)

    # Save locally
    out_name = f"internal_announcements_gold_{datetime.utcnow():%Y-%m}.xlsx"
    local_path = f"/tmp/{out_name}"
    wb.save(local_path)

    # Upload to GCS (overwrite)
    gcs_out = f"{GCS_OUTPUT_PREFIX}/{datetime.utcnow():%Y/%m}/{out_name}"
    upload_to_gcs(GCS_BUCKET, gcs_out, local_path)
    print("Uploaded", gcs_out)

if __name__ == "__main__":
    main()
