from __future__ import annotations
# ==============================================================================
# Author : Vamsi Krishna S. (enterprise build)
# Program: Territory Checks – Weekly Brand × Broker Counter (Sun→Sat)
# Build  : 3.3 (Enterprise review fixes)
#
# REVIEW NOTES
# 1) BigQuery load uses explicit schema (autodetect=False) for all 3 tables
# 2) Broker-master .xls requires xlrd installed in runtime
#    requirements.txt must include: xlrd>=2.0.1
#    If master files are converted to .xlsx, switch engine to openpyxl and update extension
# 3) Full body fetch is attempted at most once per message
# 4) Audit includes AttemptedFullBodyFetch and FetchedFullBody
# 5) Folder discovery tracks seen across recursion to avoid duplicates
# ==============================================================================

import os
import re
import time
import json
import logging
import unicodedata
import html
import csv
import argparse
from io import BytesIO

from datetime import datetime, timedelta, timezone, date
from typing import Dict, Iterable, List, Tuple, Optional

import requests
import msal
import pandas as pd
from openpyxl import Workbook, load_workbook
from google.cloud import storage, bigquery

# =============================== Configuration ================================
gv_TENANT_ID      = os.getenv("GRAPH_TENANT_ID", "")
gv_CLIENT_ID      = os.getenv("GRAPH_CLIENT_ID", "")
gv_CLIENT_SECRET  = os.getenv("GRAPH_CLIENT_SECRET", "")
gv_MASTER_MAILBOX = os.getenv("MASTER_MAILBOX", "territorycheck@strategicfranchising.com")
gv_OUT_DIR        = os.getenv("OUT_DIR", "/tmp")
gv_LOG_LEVEL      = os.getenv("LOG_LEVEL", "INFO").upper().strip()

gv_GCS_BUCKET       = os.getenv("GCS_BUCKET", "")
gv_GCS_PREFIX_EXCEL = os.getenv("GCS_PREFIX_EXCEL", "territory-checks/weekly")
gv_GCS_PREFIX_RAW   = os.getenv("GCS_PREFIX_RAW", "territory-checks/raw")

gv_GCS_BROKER_MASTER_PREFIX = os.getenv(
    "GCS_PREFIX_BROKER_MASTER",
    "territory-checks/broker-master",
)

gv_BQ_PROJECT        = os.getenv("BQ_PROJECT", "sfs-data-lake")
gv_BQ_DATASET_BRONZE = os.getenv("BQ_DATASET_BRONZE", "bronze")

gv_BQ_TABLE_TERRITORY = os.getenv("BQ_TABLE_TERRITORY", "territory_checks_raw")
gv_BQ_TABLE_UNMAPPED  = os.getenv("BQ_TABLE_UNMAPPED", "unmapped_territory_checks")
gv_BQ_TABLE_AUDIT     = os.getenv("BQ_TABLE_AUDIT", "audit_territory_checks")

gv_COUNT_FORWARDS          = True
gv_SKIP_REPLIES            = True
gv_SHOW_SUBJECT_IN_DETAILS = True

gv_USER_AGENT   = "TerritoryChecks/3.3"
gv_HTTP_TIMEOUT = 60
gv_MAX_RETRIES  = 5

logging.basicConfig(
    level=getattr(logging, gv_LOG_LEVEL, logging.INFO),
    format="%(asctime)s | %(levelname)-8s | %(message)s",
)
gv_LOG = logging.getLogger("territory-checks")

# ================================ Brand Rules =================================
gv_BRAND_ADDRESSES: Dict[str, List[str]] = {
    "Caring Transitions": ["territorycheck@caringtransitions.com", "caringtransitions.com"],
    "Fresh Coat"        : ["territorycheck@freshcoatpainters.com", "freshcoatpainters.com", "freshcoat.com"],
    "TruBlue"           : ["territorycheck@trublueally.com", "trublueally.com"],
    "Growth Coach"      : ["territorycheck@thegrowthcoach.com", "thegrowthcoach.com", "growthcoach.com"],
    "Pet Wants"         : ["territorycheck@petwants.com", "petwants.com"],
}

# ============================== Broker Mapping ================================
gv_BROKER_SUBJECT_KEYWORDS: Dict[str, str] = {
    "FranServe"        : r"\bfran\s*serve|franservesupport|franserve\b",
    "IFPG"             : r"\bifpg\b",
    "BAI"              : r"\bb\.?a\.?i\b|business\W*alliance",
    "TES"              : r"\btes\b|the\s*entrepreneur.?s\s*source|franchisesource|esourcecoach",
    "FranNet"          : r"\bfrannet\b",
    "FBA"              : r"\bfba\b|franchiseba|franchiseiba|fbamembers",
    "TPF"              : r"\btpf\b|the\s*perfect\s*franchise",
    "FCC"              : r"\bfcc\b|franchise\s*consulting\s*company",
    "Franchise Empire" : r"\bfranchise\s*empire\b",
    "SFA Advisors"     : r"\bsfa\s*advisors\b|\bsuccess\s*franchise\s*advisors\b|\bsuccess\s*fran\b",
}

gv_BROKER_DOMAIN_MAP: Dict[str, str] = {
    "franserve.com"                     : "FranServe",
    "franservesupport.com"              : "FranServe",
    "focusonfranchising.com"            : "FranServe",
    "ifpg.org"                          : "IFPG",
    "ifpg.com"                          : "IFPG",
    "businessallianceinc.com"           : "BAI",
    "zizefranchise.com"                 : "BAI",
    "inspirefranchiseconsulting.com"    : "BAI",
    "franchise-connector.com"           : "BAI",
    "franchiseconnector.com"            : "BAI",
    "markfranchise.com"                 : "BAI",
    "wwfranchiseconsulting.com"         : "BAI",
    "franchisesource.com"               : "TES",
    "esourcecoach.com"                  : "TES",
    "frannet.com"                       : "FranNet",
    "franchiseba.com"                   : "FBA",
    "franchiseiba.com"                  : "FBA",
    "fbamembers.com"                    : "FBA",
    "theperfectfranchise.com"           : "TPF",
    "thefranchiseconsultingcompany.com" : "FCC",
    "franchiseempire.com"               : "Franchise Empire",
    "sfaadvisors.com"                   : "SFA Advisors",
    "successfran.net"                   : "SFA Advisors",
    "securefranchise.com"               : "SFA Advisors",
    "successfranchiseadvisors.com"      : "SFA Advisors",
    "franocity.com"                     : "IFPG",
}

gv_BROKER_ORDER = [
    "IFPG",
    "FranServe",
    "BAI",
    "TES",
    "FranNet",
    "FBA",
    "TPF",
    "FCC",
    "Franchise Empire",
    "SFA Advisors",
    "Others",
]

# ============================== Date/Time Window ==============================
def last_completed_week_utc(lv_now_utc: Optional[datetime] = None) -> Tuple[datetime, datetime, datetime, str, str]:
    lv_now = lv_now_utc or datetime.now(timezone.utc).replace(microsecond=0)
    lv_days_since_sun = (lv_now.weekday() + 1) % 7
    lv_this_sun = (lv_now - timedelta(days=lv_days_since_sun)).replace(hour=0, minute=0, second=0, microsecond=0)
    lv_prior_sun = lv_this_sun - timedelta(days=7)

    lv_start = lv_prior_sun
    lv_end_exclusive = lv_this_sun
    lv_end_inclusive = lv_this_sun - timedelta(seconds=1)

    lv_week_label = f"{lv_start:%m.%d}-{lv_end_inclusive:%m.%d.%y}"
    lv_file_label = f"{lv_end_inclusive:%Y-%m-%d}"
    return lv_start, lv_end_inclusive, lv_end_exclusive, lv_week_label, lv_file_label

def compute_week_window(lv_override_week_end_str: Optional[str] = None) -> Tuple[datetime, datetime, datetime, str, str]:
    if lv_override_week_end_str:
        lv_end_sun = datetime.strptime(lv_override_week_end_str, "%Y-%m-%d").replace(
            tzinfo=timezone.utc, hour=0, minute=0, second=0, microsecond=0
        )
        lv_start = lv_end_sun - timedelta(days=7)
        lv_end_exclusive = lv_end_sun
        lv_end_inclusive = lv_end_sun - timedelta(seconds=1)

        lv_week_label = f"{lv_start:%m.%d}-{lv_end_inclusive:%m.%d.%y}"
        lv_file_label = f"{lv_end_inclusive:%Y-%m-%d}"
        return lv_start, lv_end_inclusive, lv_end_exclusive, lv_week_label, lv_file_label

    return last_completed_week_utc()

# ================================ Graph Client ================================
class GraphClient:
    authority = "https://login.microsoftonline.com/{tenant}"
    scope     = ["https://graph.microsoft.com/.default"]
    base      = "https://graph.microsoft.com/v1.0"

    def __init__(self, lv_tenant_id: str, lv_client_id: str, lv_client_secret: str, lv_mailbox: str):
        self.tenant_id     = lv_tenant_id
        self.client_id     = lv_client_id
        self.client_secret = lv_client_secret
        self.mailbox       = lv_mailbox
        self._token: Optional[str] = None

    def authenticate(self) -> str:
        lv_app = msal.ConfidentialClientApplication(
            self.client_id,
            authority=self.authority.format(tenant=self.tenant_id),
            client_credential=self.client_secret,
        )
        lv_result = lv_app.acquire_token_for_client(scopes=self.scope)
        if "access_token" not in lv_result:
            raise SystemExit("Graph token error:\n" + json.dumps(lv_result, indent=2))
        self._token = lv_result["access_token"]
        return self._token

    def _headers(self) -> dict:
        if not self._token:
            self.authenticate()
        return {"Authorization": f"Bearer {self._token}", "User-Agent": gv_USER_AGENT}

    def _get(self, lv_url: str, params: dict | None = None) -> requests.Response:
        lv_headers = self._headers()
        lv_backoff = 2
        last_text: str = ""
        for lv_attempt in range(1, gv_MAX_RETRIES + 1):
            try:
                lv_resp = requests.get(lv_url, headers=lv_headers, params=params, timeout=gv_HTTP_TIMEOUT)
                last_text = lv_resp.text or ""
            except requests.RequestException as lv_err:
                gv_LOG.warning("Network error (%d/%d): %s", lv_attempt, gv_MAX_RETRIES, lv_err)
                time.sleep(lv_backoff)
                lv_backoff = min(lv_backoff * 2, 32)
                continue

            if lv_resp.status_code in (429, 500, 502, 503, 504):
                lv_wait = int(lv_resp.headers.get("Retry-After", lv_backoff))
                gv_LOG.warning("Graph %s — waiting %ss (%d/%d)", lv_resp.status_code, lv_wait, lv_attempt, gv_MAX_RETRIES)
                time.sleep(lv_wait)
                lv_backoff = min(lv_backoff * 2, 32)
                continue

            if lv_resp.status_code == 401:
                gv_LOG.warning("401 Unauthorized — refreshing token (%d/%d)", lv_attempt, gv_MAX_RETRIES)
                self.authenticate()
                lv_headers = self._headers()
                continue

            if lv_resp.status_code == 403:
                raise SystemExit("403 Forbidden. Check App Access Policy.")

            try:
                lv_resp.raise_for_status()
                return lv_resp
            except Exception:
                if lv_attempt == gv_MAX_RETRIES:
                    gv_LOG.error("Graph final failure. status=%s body=%s", lv_resp.status_code, (last_text[:2000] or ""))
                raise

        raise SystemExit("Graph request failed after retries.")

    def _list_child_folders(self, lv_parent_id: str | None, seen: set[str]) -> List[dict]:
        lv_url = f"{self.base}/users/{self.mailbox}/mailFolders" + (f"/{lv_parent_id}/childFolders" if lv_parent_id else "")
        lv_params = {"$top": 50, "$select": "id,displayName,childFolderCount"}
        lv_out: List[dict] = []

        while True:
            lv_data = self._get(lv_url, params=lv_params).json()
            for f in lv_data.get("value", []):
                fid = f.get("id", "")
                if fid and fid not in seen:
                    seen.add(fid)
                    lv_out.append(f)
            lv_next = lv_data.get("@odata.nextLink")
            if not lv_next:
                break
            lv_url, lv_params = lv_next, None

        lv_final: List[dict] = []
        for lv_folder in lv_out:
            lv_final.append(lv_folder)
            if lv_folder.get("childFolderCount", 0) > 0:
                lv_final.extend(self._list_child_folders(lv_folder["id"], seen))
        return lv_final

    def list_all_folders(self) -> List[dict]:
        seen: set[str] = set()
        lv_folders = self._list_child_folders(None, seen)
        gv_LOG.info("Discovered %d mail folders.", len(lv_folders))
        return lv_folders

    def fetch_messages_in_folder(self, lv_folder_id: str, lv_start_iso: str, lv_end_iso: str) -> Iterable[dict]:
        lv_url = f"{self.base}/users/{self.mailbox}/mailFolders/{lv_folder_id}/messages"
        lv_params = {
            "$filter": f"receivedDateTime ge {lv_start_iso} and receivedDateTime lt {lv_end_iso}",
            "$orderby": "receivedDateTime asc",
            "$top": 50,
            "$select": (
                "id,internetMessageId,conversationId,receivedDateTime,subject,"
                "from,toRecipients,ccRecipients,bccRecipients,bodyPreview,internetMessageHeaders"
            ),
        }

        while True:
            lv_data = self._get(lv_url, params=lv_params).json()
            for lv_msg in lv_data.get("value", []):
                yield lv_msg
            lv_next = lv_data.get("@odata.nextLink")
            if not lv_next:
                break
            lv_url, lv_params = lv_next, None

    def fetch_full_body(self, lv_message_id: str) -> str:
        lv_url = f"{self.base}/users/{self.mailbox}/messages/{lv_message_id}"
        lv_params = {"$select": "body"}
        lv_data = self._get(lv_url, params=lv_params).json()
        return (((lv_data.get("body") or {}).get("content")) or "")

# ============================ Classification Helpers ==========================
gv_RE_PREFIX = re.compile(r"^\s*re\s*[:\-]\s*", re.IGNORECASE)
gv_FW_PREFIX = re.compile(r"^\s*(fw|fwd)\s*[:\-]\s*", re.IGNORECASE)
gv_RE_REPLY_IN_CHAT = re.compile(r"reply in chat", re.IGNORECASE)

def is_reply(lv_s: str) -> bool:
    return bool(gv_RE_PREFIX.match(lv_s or ""))

def is_forward(lv_s: str) -> bool:
    return bool(gv_FW_PREFIX.match(lv_s or ""))

def is_reply_in_chat(lv_subject: str, lv_body_preview: str) -> bool:
    return bool(gv_RE_REPLY_IN_CHAT.search(f"{lv_subject or ''} {lv_body_preview or ''}"))

def _norm(lv_text: str) -> str:
    lv_s = unicodedata.normalize("NFKC", lv_text or "").lower()
    lv_s = re.sub(r"[\u00A0\u2000-\u200B\u2060]", " ", lv_s)
    lv_s = re.sub(r"\s+", " ", lv_s).strip()
    return lv_s

def _norm_name(lv_name: str) -> str:
    lv_s = _norm(lv_name)
    lv_s = re.sub(r"[^a-z0-9 ]", " ", lv_s)
    lv_s = re.sub(r"\s+", " ", lv_s).strip()
    return lv_s

def _html_to_text(lv_html_str: str) -> str:
    if not lv_html_str:
        return ""
    lv_txt = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", lv_html_str)
    lv_txt = re.sub(r"(?is)<br\s*/?>", "\n", lv_txt)
    lv_txt = re.sub(r"(?is)</p\s*>", "\n", lv_txt)
    lv_txt = re.sub(r"(?is)<.*?>", " ", lv_txt)
    lv_txt = html.unescape(lv_txt)
    lv_txt = re.sub(r"\r", "\n", lv_txt)
    lv_txt = re.sub(r"[ \t]+", " ", lv_txt)
    lv_txt = re.sub(r"\n{3,}", "\n\n", lv_txt)
    return lv_txt.strip()

def _addr_list(*lv_lists: List[dict] | None) -> List[str]:
    lv_addrs: List[str] = []
    for lv_coll in lv_lists:
        for lv_item in (lv_coll or []):
            lv_addr = (lv_item.get("emailAddress") or {}).get("address", "")
            if lv_addr:
                lv_addrs.append(lv_addr.lower())
    return lv_addrs

def _match_brand_in_text(lv_haystack: str) -> Optional[str]:
    lv_hay = (lv_haystack or "").lower()
    for lv_brand, lv_needles in gv_BRAND_ADDRESSES.items():
        for lv_n in lv_needles:
            lv_tok = lv_n.lower()
            if "@" in lv_tok:
                if re.search(r"\b" + re.escape(lv_tok) + r"\b", lv_hay):
                    return lv_brand
            else:
                if re.search(r"(?<![a-z0-9])" + re.escape(lv_tok) + r"\b", lv_hay):
                    return lv_brand
    return None

def brand_from_recipients(lv_to_list: List[dict] | None, lv_cc_list: List[dict] | None, lv_bcc_list: List[dict] | None) -> tuple[Optional[str], str]:
    lv_addrs = _addr_list(lv_to_list, lv_cc_list, lv_bcc_list)
    lv_brand = _match_brand_in_text(" ".join(lv_addrs))
    return lv_brand, ("To/Cc/Bcc" if lv_brand else "")

gv_ROUTING_HEADERS = {"to", "cc", "bcc", "delivered-to", "x-original-to", "x-envelope-to", "x-forwarded-to", "return-path"}

def brand_from_headers(lv_headers: List[dict] | None) -> tuple[Optional[str], str]:
    if not lv_headers:
        return (None, "")
    lv_vals: List[str] = []
    for lv_h in lv_headers:
        lv_name = (lv_h.get("name") or "").lower()
        if lv_name in gv_ROUTING_HEADERS:
            lv_vals.append(lv_h.get("value", ""))
    if not lv_vals:
        return (None, "")
    lv_brand = _match_brand_in_text(" ".join(lv_vals))
    return lv_brand, ("Headers(" + ",".join(sorted(gv_ROUTING_HEADERS)) + ")" if lv_brand else "")

def broker_from_subject(lv_subject: str) -> Optional[str]:
    lv_s = _norm(lv_subject)
    for lv_broker, lv_pattern in gv_BROKER_SUBJECT_KEYWORDS.items():
        if re.search(lv_pattern, lv_s):
            return lv_broker
    return None

def broker_from_sender(lv_from_obj: dict | None) -> str:
    lv_addr = ((lv_from_obj or {}).get("emailAddress") or {}).get("address", "") or ""
    lv_dom = lv_addr.split("@")[-1].lower() if "@" in lv_addr else lv_addr.lower()
    for lv_d, lv_canonical in gv_BROKER_DOMAIN_MAP.items():
        if lv_d in lv_dom:
            return lv_canonical
    return "Others"

gv_RE_FWD_FROM_ANGLE = re.compile(
    r"^from:\s*[^<\r\n]*<\s*([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})\s*>",
    re.IGNORECASE | re.MULTILINE,
)
gv_RE_FWD_FROM_BARE = re.compile(
    r"^from:\s*([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})",
    re.IGNORECASE | re.MULTILINE,
)

def forwarded_broker_from_text(lv_text: str) -> Optional[str]:
    if not lv_text:
        return None

    lv_m = gv_RE_FWD_FROM_ANGLE.search(lv_text) or gv_RE_FWD_FROM_BARE.search(lv_text)
    if lv_m:
        lv_email = lv_m.group(1).strip().lower()
        lv_dom = lv_email.split("@")[-1]
        for lv_d, lv_canonical in gv_BROKER_DOMAIN_MAP.items():
            if lv_d in lv_dom:
                return lv_canonical

    for lv_broker, lv_pattern in gv_BROKER_SUBJECT_KEYWORDS.items():
        if re.search(lv_pattern, (lv_text or "").lower()):
            return lv_broker

    return None

# Candidate extraction for broker-master
gv_RE_ANY_EMAIL = re.compile(r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", re.IGNORECASE)
gv_RE_FROM_LINE_EMAIL = re.compile(
    r"^from:\s*(.*?)<\s*([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})\s*>",
    re.IGNORECASE | re.MULTILINE,
)

gv_FORWARD_CONTEXT_LINE = re.compile(
    r"^\s*(from:|sent:|to:|cc:|subject:|original message|forwarded message|wrote:)",
    re.IGNORECASE,
)

def extract_forwarded_from_candidates(lv_text: str, max_fallback_emails: int = 6) -> List[Tuple[str, str]]:
    if not lv_text:
        return []

    out: List[Tuple[str, str]] = []

    for m in gv_RE_FROM_LINE_EMAIL.finditer(lv_text):
        name = (m.group(1) or "").strip().strip("'\"")
        email = (m.group(2) or "").strip().lower()
        if email:
            out.append((email, name))

    if not out:
        lines = (lv_text or "").splitlines()
        for ln in lines:
            if not gv_FORWARD_CONTEXT_LINE.search(ln):
                continue
            for e in gv_RE_ANY_EMAIL.findall(ln):
                out.append((e.strip().lower(), ""))
                if len(out) >= max_fallback_emails:
                    break
            if len(out) >= max_fallback_emails:
                break

    seen = set()
    final: List[Tuple[str, str]] = []
    for email, name in out:
        if email and email not in seen:
            seen.add(email)
            final.append((email, name))
    return final

# Territory matchers
gv_TERR_PATTERNS = (
    re.compile(r"\bin\s+([a-z .'\-]+,\s*[a-z]{2})\b", re.IGNORECASE),
    re.compile(r"\b([a-z .'\-]+,\s*[a-z]{2})\b", re.IGNORECASE),
)
gv_BODY_TERR_PATTERNS = (
    re.compile(r"desired\s*territor(?:y|ies)\s*:\s*([^\r\n]+)", re.IGNORECASE),
    re.compile(r"territor(?:y|ies)\s*requested\s*:\s*([^\r\n]+)", re.IGNORECASE),
    re.compile(r"territor(?:y|ies)\s*:\s*([^\r\n]+)", re.IGNORECASE),
    re.compile(r"location\s*:\s*([^\r\n]+)", re.IGNORECASE),
)
gv_ANYWHERE_TERR_FALLBACK = re.compile(
    r"territor(?:y|ies)[^A-Za-z]{0,20}([A-Za-z .'\-]+,\s*[A-Za-z]{2})",
    re.IGNORECASE | re.DOTALL,
)

def territory_from_subject(lv_subject: str) -> str:
    lv_s = _norm(lv_subject)
    for lv_pat in gv_TERR_PATTERNS:
        lv_m = lv_pat.search(lv_s)
        if lv_m:
            return lv_m.group(1).title()
    return ""

def _clean_territory(lv_raw: str) -> str:
    lv_s = _norm(lv_raw or "")
    lv_s = re.split(r"[|/;\n\r]", lv_s)[0].strip()
    lv_s = re.sub(r"\s+", " ", lv_s)
    return lv_s.title()

def territory_from_any(lv_subject: str, lv_body_preview: str, lv_full_body_text: str | None = None) -> str:
    lv_terr = territory_from_subject(lv_subject)
    if lv_terr:
        return lv_terr

    for lv_pat in gv_BODY_TERR_PATTERNS:
        lv_m = lv_pat.search(lv_body_preview or "")
        if lv_m:
            return _clean_territory(lv_m.group(1))

    lv_m = gv_ANYWHERE_TERR_FALLBACK.search(lv_body_preview or "")
    if lv_m:
        return _clean_territory(lv_m.group(1))

    if lv_full_body_text:
        for lv_pat in gv_BODY_TERR_PATTERNS:
            lv_m = lv_pat.search(lv_full_body_text)
            if lv_m:
                return _clean_territory(lv_m.group(1))
        lv_m = gv_ANYWHERE_TERR_FALLBACK.search(lv_full_body_text)
        if lv_m:
            return _clean_territory(lv_m.group(1))
        lv_m = re.search(r"\b([A-Za-z .'\-]+,\s*[A-Za-z]{2})\b", lv_full_body_text)
        if lv_m:
            return _clean_territory(lv_m.group(1))

    return ""

def pretty_label_from_domain(lv_email_or_domain: str) -> str:
    lv_dom = (lv_email_or_domain or "").split("@")[-1].lower()
    if not lv_dom or "." not in lv_dom:
        return "Unknown"
    lv_core = lv_dom.split(".")[0]
    lv_chunks = re.split(r"[\W_]+", lv_core)
    if not lv_chunks or lv_chunks == [""]:
        return lv_core.title()
    return " ".join([lv_c.upper() if len(lv_c) <= 3 else lv_c.title() for lv_c in lv_chunks if lv_c])

# ================================ Excel Helpers ===============================
def ensure_out_dir() -> None:
    os.makedirs(gv_OUT_DIR, exist_ok=True)

def unique_week_filepath(lv_end_label: str) -> str:
    lv_base = os.path.join(gv_OUT_DIR, f"Territory_Checks_{lv_end_label}.xlsx")
    if not os.path.exists(lv_base):
        return lv_base
    lv_i = 1
    while True:
        lv_cand = os.path.join(gv_OUT_DIR, f"Territory_Checks_{lv_end_label} ({lv_i}).xlsx")
        if not os.path.exists(lv_cand):
            return lv_cand
        lv_i += 1

def try_open_workbook(lv_path: str) -> Optional[Workbook]:
    if not os.path.exists(lv_path):
        lv_wb = Workbook()
        lv_wb.active.title = "Index"
        return lv_wb
    try:
        return load_workbook(lv_path)
    except PermissionError:
        return None

# ============================ BigQuery helpers ================================
def bq_client() -> bigquery.Client:
    return bigquery.Client(project=gv_BQ_PROJECT) if gv_BQ_PROJECT else bigquery.Client()

def delete_week_slice(dataset: str, table: str, run_date_from: date, run_date_to: date) -> None:
    if not gv_BQ_PROJECT:
        gv_LOG.info("BQ_PROJECT not set; skipping delete-week-slice for %s.%s", dataset, table)
        return

    table_id = f"{gv_BQ_PROJECT}.{dataset}.{table}"
    client = bq_client()
    sql = f"""
        DELETE FROM `{table_id}`
        WHERE run_date_from = @from_date
          AND run_date_to   = @to_date
    """
    job_config = bigquery.QueryJobConfig(
        query_parameters=[
            bigquery.ScalarQueryParameter("from_date", "DATE", run_date_from),
            bigquery.ScalarQueryParameter("to_date",   "DATE", run_date_to),
        ]
    )
    gv_LOG.info("Deleting existing week slice from %s for run_date_from=%s, run_date_to=%s", table_id, run_date_from, run_date_to)
    res = client.query(sql, job_config=job_config).result()
    gv_LOG.info("Deleted %s rows from %s for the given week slice.", res.num_dml_affected_rows, table_id)

# ============================ Path helper for GCS =============================
def join_prefix(base_prefix: str, *parts: str) -> str:
    base = (base_prefix or "").strip().strip("/")
    extra_parts = [p.strip().strip("/") for p in parts if p]
    extra = "/".join(extra_parts)
    if base and extra:
        return f"{base}/{extra}"
    return extra or base

# ======================= Broker master (brand-only fallback) ==================
gv_BROKER_MASTER_CONFIG = {
    "Caring Transitions": {"slug": "caring_transitions", "header_row": 2, "sheet": 0},
    "Fresh Coat":         {"slug": "freshcoat",         "header_row": 2, "sheet": 0},
    "Pet Wants":          {"slug": "petwants",          "header_row": 2, "sheet": 0},
    "TruBlue":            {"slug": "trublue",           "header_row": 2, "sheet": 0},
}

gv_BROKER_MASTER_CACHE: Dict[str, Dict[str, Dict[str, str]]] = {}

def _canonical_agency(lv_agency: str) -> Optional[str]:
    if not lv_agency:
        return None
    a = _norm(lv_agency)

    alias = {
        "franserve": "FranServe",
        "fran serve": "FranServe",
        "ifpg": "IFPG",
        "bai": "BAI",
        "business alliance": "BAI",
        "the entrepreneur s source": "TES",
        "entrepreneur s source": "TES",
        "entrepreneur's source": "TES",
        "entrepreneur source": "TES",
        "tes": "TES",
        "frannet": "FranNet",
        "fba": "FBA",
        "the perfect franchise": "TPF",
        "tpf": "TPF",
        "fcc": "FCC",
        "franchise consulting company": "FCC",
        "franchise empire": "Franchise Empire",
        "sfa advisors": "SFA Advisors",
        "success franchise advisors": "SFA Advisors",
    }

    if a in alias:
        return alias[a]
    if "franserve" in a:
        return "FranServe"
    if "ifpg" in a:
        return "IFPG"
    if "business" in a and "alliance" in a:
        return "BAI"
    if "entrepreneur" in a and ("source" in a or "esource" in a):
        return "TES"

    for b in gv_BROKER_ORDER:
        if _norm(b) == a:
            return b

    return None

def load_broker_master_for_brand(brand_name: str) -> Dict[str, Dict[str, str]]:
    if brand_name in gv_BROKER_MASTER_CACHE:
        return gv_BROKER_MASTER_CACHE[brand_name]

    cfg = gv_BROKER_MASTER_CONFIG.get(brand_name)
    if not cfg:
        gv_BROKER_MASTER_CACHE[brand_name] = {"email_map": {}, "name_map": {}}
        return gv_BROKER_MASTER_CACHE[brand_name]

    if not gv_GCS_BUCKET:
        gv_LOG.warning("GCS_BUCKET not set; skipping broker master for brand %s", brand_name)
        gv_BROKER_MASTER_CACHE[brand_name] = {"email_map": {}, "name_map": {}}
        return gv_BROKER_MASTER_CACHE[brand_name]

    blob_path = join_prefix(gv_GCS_BROKER_MASTER_PREFIX, f"broker_master_{cfg['slug']}.xls")

    client = storage.Client()
    bucket = client.bucket(gv_GCS_BUCKET)
    blob = bucket.blob(blob_path)

    if not blob.exists(client=client):
        gv_LOG.warning("Broker master file not found: gs://%s/%s", gv_GCS_BUCKET, blob_path)
        gv_BROKER_MASTER_CACHE[brand_name] = {"email_map": {}, "name_map": {}}
        return gv_BROKER_MASTER_CACHE[brand_name]

    buf = BytesIO(blob.download_as_bytes())
    buf.seek(0)

    try:
        df = pd.read_excel(
            buf,
            engine="xlrd",
            sheet_name=cfg["sheet"],
            header=cfg["header_row"],
        )
    except Exception as e:
        gv_LOG.warning("Error reading broker master for brand %s at %s: %s", brand_name, blob_path, e)
        gv_BROKER_MASTER_CACHE[brand_name] = {"email_map": {}, "name_map": {}}
        return gv_BROKER_MASTER_CACHE[brand_name]

    df.columns = [str(c).strip() for c in df.columns]

    name_col = "Broker's Name"
    agency_col = "Agency"
    email_col = "Email"

    if agency_col not in df.columns:
        gv_LOG.warning("Broker master for brand %s missing Agency column. Columns=%s", brand_name, list(df.columns))
        gv_BROKER_MASTER_CACHE[brand_name] = {"email_map": {}, "name_map": {}}
        return gv_BROKER_MASTER_CACHE[brand_name]

    email_map: Dict[str, str] = {}
    name_map: Dict[str, str] = {}

    for _, r in df.iterrows():
        agency_raw = str(r.get(agency_col, "")).strip()
        agency = _canonical_agency(agency_raw)
        if not agency:
            continue

        email = str(r.get(email_col, "")).strip().lower() if email_col in df.columns else ""
        broker_name = str(r.get(name_col, "")).strip() if name_col in df.columns else ""

        if email and "@" in email:
            email_map[email] = agency

        nn = _norm_name(broker_name)
        if nn:
            name_map[nn] = agency

    gv_LOG.info("Loaded broker master for %s: %d emails, %d names", brand_name, len(email_map), len(name_map))
    gv_BROKER_MASTER_CACHE[brand_name] = {"email_map": email_map, "name_map": name_map}
    return gv_BROKER_MASTER_CACHE[brand_name]

def resolve_broker_from_master_candidates(brand_name: str, candidates: List[Tuple[str, str]]) -> Optional[str]:
    maps = load_broker_master_for_brand(brand_name)
    email_map = maps.get("email_map", {})
    name_map  = maps.get("name_map", {})

    for email, _name in (candidates or []):
        e = (email or "").strip().lower()
        if e and e in email_map:
            return email_map[e]

    for _email, name in (candidates or []):
        sn = _norm_name(name or "")
        if not sn:
            continue
        if sn in name_map:
            return name_map[sn]
        for nm, ag in name_map.items():
            if nm and (nm in sn or sn in nm):
                return ag

    return None

# ================================== Main ======================================
def run(lv_override_week_end_str: Optional[str] = None) -> dict:
    lv_missing = [
        lv_k
        for lv_k, lv_v in {
            "GRAPH_TENANT_ID"    : gv_TENANT_ID,
            "GRAPH_CLIENT_ID"    : gv_CLIENT_ID,
            "GRAPH_CLIENT_SECRET": gv_CLIENT_SECRET,
            "MASTER_MAILBOX"     : gv_MASTER_MAILBOX,
            "OUT_DIR"            : gv_OUT_DIR,
        }.items()
        if not lv_v
    ]
    if lv_missing:
        raise SystemExit("Missing environment variables: " + ", ".join(lv_missing))

    ensure_out_dir()

    lv_start_dt, lv_end_dt, lv_end_excl, lv_week_label, lv_file_label = compute_week_window(lv_override_week_end_str)
    lv_run_ts = datetime.now(timezone.utc).replace(microsecond=0)

    lv_run_date_from = lv_start_dt.date()
    lv_run_date_to   = lv_end_dt.date()

    lv_year  = f"{lv_end_dt.year:04d}"
    lv_month = f"{lv_end_dt.month:02d}"

    gv_LOG.info("Processing week window: %s → %s (%s) [end_exclusive=%s]", lv_start_dt, lv_end_dt, lv_week_label, lv_end_excl)

    lv_client = GraphClient(gv_TENANT_ID, gv_CLIENT_ID, gv_CLIENT_SECRET, gv_MASTER_MAILBOX)
    lv_client.authenticate()

    lv_start_iso = lv_start_dt.strftime("%Y-%m-%dT%H:%M:%SZ")
    lv_end_iso   = lv_end_excl.strftime("%Y-%m-%dT%H:%M:%SZ")

    lv_details_rows: List[dict] = []
    lv_audit_rows:   List[dict] = []

    for lv_folder in lv_client.list_all_folders():
        lv_fid   = lv_folder["id"]
        lv_fname = lv_folder.get("displayName", "")
        gv_LOG.info("Scanning folder: %s", lv_fname)

        for lv_msg in lv_client.fetch_messages_in_folder(lv_fid, lv_start_iso, lv_end_iso):
            lv_subj = (lv_msg.get("subject") or "").strip()
            lv_is_reply   = is_reply(lv_subj)
            lv_is_forward = is_forward(lv_subj)
            lv_body_preview = (lv_msg.get("bodyPreview") or "").strip()

            lv_full_text: Optional[str] = None
            lv_attempted_full_body_fetch: bool = False
            lv_fetched_full_body: bool = False

            lv_skipped_reason = ""
            if is_reply_in_chat(lv_subj, lv_body_preview):
                lv_skipped_reason = "Reply in Chat"
            elif gv_SKIP_REPLIES and lv_is_reply:
                lv_skipped_reason = "Reply"
            elif (not gv_COUNT_FORWARDS) and lv_is_forward:
                lv_skipped_reason = "Forward (disabled)"

            lv_b_to_cc_bcc, lv_source1 = brand_from_recipients(
                lv_msg.get("toRecipients"),
                lv_msg.get("ccRecipients"),
                lv_msg.get("bccRecipients"),
            )
            lv_b_hdrs, lv_source2 = (None, "")
            if not lv_b_to_cc_bcc:
                lv_b_hdrs, lv_source2 = brand_from_headers(lv_msg.get("internetMessageHeaders"))

            lv_brand        = lv_b_to_cc_bcc or lv_b_hdrs or "Unmapped"
            lv_brand_source = lv_source1 or lv_source2 or "Unmapped"

            lv_from_obj    = (lv_msg.get("from") or {}).get("emailAddress") or {}
            lv_sender      = (lv_from_obj.get("address") or "").strip()
            lv_sender_name = (lv_from_obj.get("name") or "").strip() or lv_sender

            lv_broker_subj = broker_from_subject(lv_subj)
            lv_broker_fw   = None if lv_broker_subj else forwarded_broker_from_text(lv_body_preview)
            lv_broker_dom  = broker_from_sender(lv_msg.get("from"))
            lv_chosen_broker = lv_broker_subj or lv_broker_fw or lv_broker_dom

            def _ensure_full_text_once() -> None:
                nonlocal lv_full_text, lv_attempted_full_body_fetch, lv_fetched_full_body
                if lv_attempted_full_body_fetch:
                    return
                lv_attempted_full_body_fetch = True
                try:
                    lv_body_html = lv_client.fetch_full_body(lv_msg["id"])
                    lv_full_text = _html_to_text(lv_body_html)
                    lv_fetched_full_body = True
                except Exception as lv_e:
                    gv_LOG.debug("Full body fetch failed for %s: %s", lv_msg.get("id"), lv_e)
                    lv_full_text = None
                    lv_fetched_full_body = False

            if lv_chosen_broker == "Others":
                _ensure_full_text_once()
                lv_broker_full = forwarded_broker_from_text(lv_full_text or "")
                if lv_broker_full:
                    lv_chosen_broker = lv_broker_full

            if lv_chosen_broker == "Others" and lv_brand in gv_BROKER_MASTER_CONFIG:
                _ensure_full_text_once()
                candidates = extract_forwarded_from_candidates(lv_full_text or "")
                if lv_sender and "@" in lv_sender:
                    candidates.append((lv_sender.strip().lower(), lv_sender_name))

                lv_master = resolve_broker_from_master_candidates(lv_brand, candidates)
                if lv_master:
                    gv_LOG.info("Broker-master resolved: %s -> %s (brand=%s)", lv_sender, lv_master, lv_brand)
                    lv_chosen_broker = lv_master

            lv_detail_broker_label = lv_chosen_broker
            lv_others_label: Optional[str] = None
            if lv_chosen_broker == "Others":
                lv_pretty = pretty_label_from_domain(lv_sender)
                lv_detail_broker_label = f"Others | {lv_pretty}"
                lv_others_label = lv_pretty

            lv_terr = territory_from_any(lv_subj, lv_body_preview, lv_full_text)
            if not lv_terr:
                _ensure_full_text_once()
                lv_terr = territory_from_any(lv_subj, lv_body_preview, lv_full_text)

            lv_recv    = lv_msg.get("receivedDateTime", "")
            lv_recv_dt = lv_recv[:10] if lv_recv else ""

            lv_to_str  = ";".join([(x.get("emailAddress") or {}).get("address", "") for x in (lv_msg.get("toRecipients") or [])])
            lv_cc_str  = ";".join([(x.get("emailAddress") or {}).get("address", "") for x in (lv_msg.get("ccRecipients") or [])])
            lv_bcc_str = ";".join([(x.get("emailAddress") or {}).get("address", "") for x in (lv_msg.get("bccRecipients") or [])])

            lv_audit_rows.append({
                "Folder"                : lv_fname,
                "Brand"                 : lv_brand,
                "Brand Source"          : lv_brand_source,
                "IsForward"             : lv_is_forward,
                "IsReply"               : lv_is_reply,
                "Subject"               : lv_subj,
                "From"                  : lv_sender,
                "To"                    : lv_to_str,
                "CC"                    : lv_cc_str,
                "BCC"                   : lv_bcc_str,
                "BodyPreview"           : lv_body_preview,
                "ReceivedUTC"           : lv_recv,
                "FetchedFullBody"       : lv_fetched_full_body,
                "AttemptedFullBodyFetch": lv_attempted_full_body_fetch,
                "Chosen Broker (bucket)": lv_chosen_broker,
                "SkippedReason"         : lv_skipped_reason,
            })

            if lv_skipped_reason:
                continue

            lv_details_rows.append({
                "Brand"              : lv_brand,
                "Broker Brand"       : lv_detail_broker_label,
                "Summary Bucket"     : lv_chosen_broker,
                "Others Name"        : lv_others_label or "",
                "Broker Name"        : lv_sender_name,
                "Territory"          : lv_terr,
                "Received (UTC Date)": lv_recv_dt,
                "Subject"            : lv_subj,
                "From"               : lv_sender,
                "Folder"             : lv_fname,
                "Brand Source"       : lv_brand_source,
                "ConversationId"     : lv_msg.get("conversationId", ""),
                "InternetMessageId"  : lv_msg.get("internetMessageId", ""),
            })

    lv_df = pd.DataFrame(lv_details_rows)

    if lv_df.empty:
        gv_LOG.info("No messages in the window; writing an empty workbook with all brand sheets.")
        empty_df = pd.DataFrame(columns=[
            "Brand","Broker Brand","Summary Bucket","Others Name","Broker Name","Territory",
            "Received (UTC Date)","Subject","From","Folder","Brand Source","ConversationId","InternetMessageId"
        ])
        excel_local_path = _write_workbook(empty_df, lv_audit_rows, lv_week_label, lv_file_label)
        excel_prefix = join_prefix(gv_GCS_PREFIX_EXCEL, lv_year, lv_month)
        excel_gcs_uri = upload_file_to_gcs(excel_local_path, excel_prefix)
        return {"excel_local_path": excel_local_path, "excel_gcs_uri": excel_gcs_uri, "raw_local_path": None, "raw_gcs_uri": None}

    if "InternetMessageId" in lv_df.columns:
        lv_df = lv_df.sort_values(by=["InternetMessageId", "Received (UTC Date)"])
        lv_df = lv_df.drop_duplicates(subset=["InternetMessageId"], keep="first")

    def _canon_terr(lv_x: str) -> str:
        lv_s = _norm(lv_x or "")
        lv_s = re.sub(r"[^a-z0-9 ,.'\-]", " ", lv_s)
        lv_s = re.sub(r"\s+", " ", lv_s).strip()
        return lv_s

    lv_df["__TerrKey__"] = lv_df["Territory"].map(_canon_terr)
    lv_with_terr    = lv_df[lv_df["__TerrKey__"] != ""]
    lv_without_terr = lv_df[lv_df["__TerrKey__"] == ""]
    lv_deduped = lv_with_terr.drop_duplicates(subset=["Brand", "Summary Bucket", "Broker Brand", "__TerrKey__"], keep="first")
    lv_df = pd.concat([lv_deduped, lv_without_terr], ignore_index=True)

    lv_df["Received (UTC Date)"] = pd.to_datetime(lv_df["Received (UTC Date)"], errors="coerce")
    lv_df.sort_values(by=["Brand", "Summary Bucket", "Received (UTC Date)", "Subject"], inplace=True)
    lv_df["Seq"] = lv_df.groupby(["Brand", "Broker Brand"]).cumcount().add(1)

    df_territory = lv_df.copy()
    df_territory["run_date_from"] = lv_run_date_from
    df_territory["run_date_to"]   = lv_run_date_to
    df_territory["run_timestamp"] = lv_run_ts

    df_territory = df_territory.rename(columns={
        "Brand": "brand_name",
        "Broker Brand": "broker_brand",
        "Broker Name": "broker_name",
        "Territory": "territory",
        "Received (UTC Date)": "received_utc",
        "Seq": "seq",
        "Subject": "subject",
    })

    df_territory = df_territory[[
        "brand_name","broker_brand","broker_name","territory","received_utc","seq","subject",
        "run_date_from","run_date_to","run_timestamp"
    ]]

    df_unmapped = lv_df[lv_df["Brand"] == "Unmapped"].copy()
    df_unmapped["run_date_from"] = lv_run_date_from
    df_unmapped["run_date_to"]   = lv_run_date_to
    df_unmapped["run_timestamp"] = lv_run_ts

    df_unmapped = df_unmapped.rename(columns={
        "Brand": "brand_name",
        "Brand Source": "brand_source",
        "Broker Brand": "broker_brand",
        "Broker Name": "broker_name",
        "Territory": "territory",
        "Received (UTC Date)": "received_utc",
        "Subject": "subject",
        "From": "from_email",
        "Folder": "folder_name",
    })

    df_unmapped = df_unmapped[[
        "brand_name","brand_source","broker_brand","broker_name","territory","received_utc",
        "subject","from_email","folder_name","run_date_from","run_date_to","run_timestamp"
    ]]

    df_audit = pd.DataFrame(lv_audit_rows)
    if not df_audit.empty:
        df_audit["run_date_from"] = lv_run_date_from
        df_audit["run_date_to"]   = lv_run_date_to
        df_audit["run_timestamp"] = lv_run_ts

        df_audit = df_audit.rename(columns={
            "Folder": "folder_name",
            "Brand": "brand_name",
            "Brand Source": "brand_source",
            "IsForward": "is_forward",
            "IsReply": "is_reply",
            "Subject": "subject",
            "From": "from_email",
            "To": "to_email",
            "CC": "cc_email",
            "BCC": "bcc_email",
            "BodyPreview": "body_preview",
            "ReceivedUTC": "received_utc",
            "FetchedFullBody": "fetched_full_body",
            "AttemptedFullBodyFetch": "attempted_full_body_fetch",
            "Chosen Broker (bucket)": "chosen_broker",
            "SkippedReason": "skipped_reason",
        })

        df_audit = df_audit[[
            "folder_name","brand_name","brand_source","is_forward","is_reply","subject","from_email",
            "to_email","cc_email","bcc_email","body_preview","received_utc","fetched_full_body",
            "attempted_full_body_fetch","chosen_broker","skipped_reason",
            "run_date_from","run_date_to","run_timestamp"
        ]]

    raw_filename      = f"territory_checks_raw_{lv_file_label}.csv"
    raw_unmapped_file = f"territory_unmapped_{lv_file_label}.csv"
    audit_filename    = f"territory_audit_{lv_file_label}.csv"

    raw_local_path      = os.path.join(gv_OUT_DIR, raw_filename)
    unmapped_local_path = os.path.join(gv_OUT_DIR, raw_unmapped_file)
    audit_local_path    = os.path.join(gv_OUT_DIR, audit_filename)

    common_date_format = "%Y-%m-%d %H:%M:%S"

    for df in (df_territory, df_unmapped, df_audit):
        if df is not None and not df.empty and "received_utc" in df.columns:
            df["received_utc"] = pd.to_datetime(df["received_utc"], errors="coerce")

    df_territory.to_csv(raw_local_path, index=False, date_format=common_date_format)
    df_unmapped.to_csv(unmapped_local_path, index=False, date_format=common_date_format)

    if df_audit is not None and not df_audit.empty:
        if "run_timestamp" in df_audit.columns:
            df_audit["run_timestamp"] = pd.to_datetime(df_audit["run_timestamp"], errors="coerce").dt.strftime(common_date_format)
        df_audit.to_csv(audit_local_path, index=False, quoting=csv.QUOTE_ALL, lineterminator="\n")

    gv_LOG.info("Saved CSVs for BigQuery: %s, %s, %s", raw_local_path, unmapped_local_path, audit_local_path)

    excel_local_path = _write_workbook(lv_df, lv_audit_rows, lv_week_label, lv_file_label)

    excel_prefix     = join_prefix(gv_GCS_PREFIX_EXCEL, lv_year, lv_month)
    territory_prefix = join_prefix(gv_GCS_PREFIX_RAW, "territory", lv_year, lv_month)
    unmapped_prefix  = join_prefix(gv_GCS_PREFIX_RAW, "unmapped", lv_year, lv_month)
    audit_prefix     = join_prefix(gv_GCS_PREFIX_RAW, "audit", lv_year, lv_month)

    raw_gcs_uri      = upload_file_to_gcs(raw_local_path,      territory_prefix)
    unmapped_gcs_uri = upload_file_to_gcs(unmapped_local_path, unmapped_prefix)
    audit_gcs_uri    = upload_file_to_gcs(audit_local_path,    audit_prefix)
    excel_gcs_uri    = upload_file_to_gcs(excel_local_path,    excel_prefix)

    if raw_gcs_uri:
        delete_week_slice(gv_BQ_DATASET_BRONZE, gv_BQ_TABLE_TERRITORY, lv_run_date_from, lv_run_date_to)
        load_csv_to_bigquery(raw_gcs_uri, gv_BQ_DATASET_BRONZE, gv_BQ_TABLE_TERRITORY)

    if unmapped_gcs_uri:
        delete_week_slice(gv_BQ_DATASET_BRONZE, gv_BQ_TABLE_UNMAPPED, lv_run_date_from, lv_run_date_to)
        load_csv_to_bigquery(unmapped_gcs_uri, gv_BQ_DATASET_BRONZE, gv_BQ_TABLE_UNMAPPED)

    if audit_gcs_uri:
        delete_week_slice(gv_BQ_DATASET_BRONZE, gv_BQ_TABLE_AUDIT, lv_run_date_from, lv_run_date_to)
        load_csv_to_bigquery(audit_gcs_uri, gv_BQ_DATASET_BRONZE, gv_BQ_TABLE_AUDIT)

    return {
        "excel_local_path": excel_local_path,
        "excel_gcs_uri"   : excel_gcs_uri,
        "raw_local_path"  : raw_local_path,
        "raw_gcs_uri"     : raw_gcs_uri,
    }

# ------------------------------- Writer ---------------------------------------
def _write_workbook(lv_df: pd.DataFrame, lv_audit_rows: List[dict], lv_week_label: str, lv_file_label: str) -> str:
    lv_seen_brokers = [
        lv_b for lv_b in lv_df.get("Summary Bucket", pd.Series([], dtype=str)).unique().tolist()
        if lv_b not in gv_BROKER_ORDER and lv_b not in (None, "Unmapped")
    ]
    lv_broker_order_full = gv_BROKER_ORDER + sorted([lv_b for lv_b in lv_seen_brokers if isinstance(lv_b, str)])

    lv_target_path = unique_week_filepath(lv_file_label)
    lv_wb = try_open_workbook(lv_target_path)
    if lv_wb is None:
        lv_i = 1
        while lv_wb is None:
            lv_alt = os.path.join(gv_OUT_DIR, f"Territory_Checks_{lv_file_label} ({lv_i}).xlsx")
            lv_wb = try_open_workbook(lv_alt)
            if lv_wb:
                lv_target_path = lv_alt
                break
            lv_i += 1

    if not lv_df.empty:
        for lv_brand, lv_sub in lv_df.groupby("Brand"):
            lv_sheet_name = lv_brand
            if lv_sheet_name not in lv_wb.sheetnames:
                lv_wb.create_sheet(lv_sheet_name)
            lv_ws = lv_wb[lv_sheet_name]
            lv_ws.delete_rows(1, lv_ws.max_row)

            lv_cols_detail = ["Broker Brand","Broker Name","Territory","Received (UTC Date)","Seq"]
            if gv_SHOW_SUBJECT_IN_DETAILS:
                lv_cols_detail.append("Subject")

            lv_sub_out = lv_sub[lv_cols_detail + ["Summary Bucket"]].copy()
            lv_sub_out["__order__"] = lv_sub_out["Summary Bucket"].apply(
                lambda lv_b: lv_broker_order_full.index(lv_b) if lv_b in lv_broker_order_full else 999
            )
            lv_sub_out.sort_values(by=["__order__", "Received (UTC Date)", "Seq"], inplace=True)
            lv_sub_out.drop(columns=["__order__", "Summary Bucket"], inplace=True)

            lv_ws.append(lv_cols_detail)
            for _, lv_r in lv_sub_out.iterrows():
                lv_ws.append([lv_r[lv_c] for lv_c in lv_cols_detail])

            lv_col_base = 7
            lv_summary = lv_sub.groupby("Summary Bucket").size().rename("Count").reset_index()
            lv_summary = lv_summary.set_index("Summary Bucket").reindex(lv_broker_order_full, fill_value=0).reset_index()
            lv_total_row = pd.DataFrame([{"Summary Bucket": "Total", "Count": int(lv_summary["Count"].sum())}])
            lv_summary = pd.concat([lv_summary, lv_total_row], ignore_index=True)

            lv_ws.cell(row=1, column=lv_col_base,   value="Broker Brand")
            lv_ws.cell(row=1, column=lv_col_base+1, value=lv_week_label)

            for lv_i, lv_row in lv_summary.iterrows():
                lv_ws.cell(row=lv_i + 2, column=lv_col_base,     value=lv_row["Summary Bucket"])
                lv_ws.cell(row=lv_i + 2, column=lv_col_base + 1, value=int(lv_row["Count"]))

    lv_cols_detail = ["Broker Brand","Broker Name","Territory","Received (UTC Date)","Seq"]
    if gv_SHOW_SUBJECT_IN_DETAILS:
        lv_cols_detail.append("Subject")

    for lv_brand_name in gv_BRAND_ADDRESSES.keys():
        if lv_brand_name not in lv_wb.sheetnames:
            lv_wb.create_sheet(lv_brand_name)
        lv_ws = lv_wb[lv_brand_name]
        if lv_ws.max_row <= 1 and lv_ws.max_column <= 1:
            lv_ws.delete_rows(1, lv_ws.max_row)
            lv_ws.append(lv_cols_detail)
            lv_col_base = 7
            lv_ws.cell(row=1, column=lv_col_base,   value="Broker Brand")
            lv_ws.cell(row=1, column=lv_col_base+1, value=lv_week_label)
            for lv_i, lv_b in enumerate(lv_broker_order_full, start=2):
                lv_ws.cell(row=lv_i, column=lv_col_base,     value=lv_b)
                lv_ws.cell(row=lv_i, column=lv_col_base + 1, value=0)
            lv_ws.cell(row=len(lv_broker_order_full) + 2, column=lv_col_base,     value="Total")
            lv_ws.cell(row=len(lv_broker_order_full) + 2, column=lv_col_base + 1, value=0)

    if "Unmapped" not in lv_wb.sheetnames:
        lv_wb.create_sheet("Unmapped")
    lv_ws_u = lv_wb["Unmapped"]
    lv_ws_u.delete_rows(1, lv_ws_u.max_row)
    lv_cols_u = ["Brand","Brand Source","Broker Brand","Broker Name","Territory","Received (UTC Date)","Subject","From","Folder"]
    lv_ws_u.append(lv_cols_u)
    if not lv_df.empty and "Brand" in lv_df.columns:
        for _, lv_r in lv_df[lv_df["Brand"] == "Unmapped"][lv_cols_u].iterrows():
            lv_ws_u.append([lv_r.get(lv_c, "") for lv_c in lv_cols_u])

    if "Audit" not in lv_wb.sheetnames:
        lv_wb.create_sheet("Audit")
    lv_ws_a = lv_wb["Audit"]
    lv_ws_a.delete_rows(1, lv_ws_a.max_row)
    lv_cols_a = [
        "Folder","Brand","Brand Source","IsForward","IsReply","Subject","From","To","CC","BCC",
        "BodyPreview","ReceivedUTC","FetchedFullBody","AttemptedFullBodyFetch","Chosen Broker (bucket)","SkippedReason",
    ]
    lv_ws_a.append(lv_cols_a)
    for lv_row in lv_audit_rows:
        lv_ws_a.append([lv_row.get(lv_c, "") for lv_c in lv_cols_a])

    if "Index" in lv_wb.sheetnames and len(lv_wb.sheetnames) > 1:
        del lv_wb["Index"]

    lv_wb.save(lv_target_path)
    gv_LOG.info("Saved weekly workbook: %s", lv_target_path)
    return lv_target_path

def upload_file_to_gcs(local_path: str, prefix: str) -> Optional[str]:
    if not gv_GCS_BUCKET:
        gv_LOG.info("GCS_BUCKET not set; skipping upload for %s", local_path)
        return None

    client = storage.Client()
    bucket = client.bucket(gv_GCS_BUCKET)

    base_name = os.path.basename(local_path)
    prefix = (prefix or "").strip().strip("/")
    blob_name = f"{prefix}/{base_name}" if prefix else base_name

    blob = bucket.blob(blob_name)
    blob.upload_from_filename(local_path)

    gcs_uri = f"gs://{gv_GCS_BUCKET}/{blob_name}"
    gv_LOG.info("Uploaded %s to %s", local_path, gcs_uri)
    return gcs_uri

def _bq_schema_for_table(table: str) -> List[bigquery.SchemaField]:
    if table == gv_BQ_TABLE_TERRITORY:
        return [
            bigquery.SchemaField("brand_name", "STRING"),
            bigquery.SchemaField("broker_brand", "STRING"),
            bigquery.SchemaField("broker_name", "STRING"),
            bigquery.SchemaField("territory", "STRING"),
            bigquery.SchemaField("received_utc", "TIMESTAMP"),
            bigquery.SchemaField("seq", "INT64"),
            bigquery.SchemaField("subject", "STRING"),
            bigquery.SchemaField("run_date_from", "DATE"),
            bigquery.SchemaField("run_date_to", "DATE"),
            bigquery.SchemaField("run_timestamp", "TIMESTAMP"),
        ]
    if table == gv_BQ_TABLE_UNMAPPED:
        return [
            bigquery.SchemaField("brand_name", "STRING"),
            bigquery.SchemaField("brand_source", "STRING"),
            bigquery.SchemaField("broker_brand", "STRING"),
            bigquery.SchemaField("broker_name", "STRING"),
            bigquery.SchemaField("territory", "STRING"),
            bigquery.SchemaField("received_utc", "TIMESTAMP"),
            bigquery.SchemaField("subject", "STRING"),
            bigquery.SchemaField("from_email", "STRING"),
            bigquery.SchemaField("folder_name", "STRING"),
            bigquery.SchemaField("run_date_from", "DATE"),
            bigquery.SchemaField("run_date_to", "DATE"),
            bigquery.SchemaField("run_timestamp", "TIMESTAMP"),
        ]
    if table == gv_BQ_TABLE_AUDIT:
        return [
            bigquery.SchemaField("folder_name", "STRING"),
            bigquery.SchemaField("brand_name", "STRING"),
            bigquery.SchemaField("brand_source", "STRING"),
            bigquery.SchemaField("is_forward", "BOOL"),
            bigquery.SchemaField("is_reply", "BOOL"),
            bigquery.SchemaField("subject", "STRING"),
            bigquery.SchemaField("from_email", "STRING"),
            bigquery.SchemaField("to_email", "STRING"),
            bigquery.SchemaField("cc_email", "STRING"),
            bigquery.SchemaField("bcc_email", "STRING"),
            bigquery.SchemaField("body_preview", "STRING"),
            bigquery.SchemaField("received_utc", "TIMESTAMP"),
            bigquery.SchemaField("fetched_full_body", "BOOL"),
            bigquery.SchemaField("attempted_full_body_fetch", "BOOL"),
            bigquery.SchemaField("chosen_broker", "STRING"),
            bigquery.SchemaField("skipped_reason", "STRING"),
            bigquery.SchemaField("run_date_from", "DATE"),
            bigquery.SchemaField("run_date_to", "DATE"),
            bigquery.SchemaField("run_timestamp", "TIMESTAMP"),
        ]
    return []

def load_csv_to_bigquery(gcs_uri: str, dataset: str, table: str) -> None:
    if not gv_BQ_PROJECT:
        gv_LOG.info("BQ_PROJECT not set; skipping BigQuery load for %s", gcs_uri)
        return

    table_id = f"{gv_BQ_PROJECT}.{dataset}.{table}"
    client = bigquery.Client(project=gv_BQ_PROJECT)

    schema = _bq_schema_for_table(table)
    if not schema:
        raise SystemExit(f"No schema configured for table={table}. Update _bq_schema_for_table().")

    job_config = bigquery.LoadJobConfig(
        source_format=bigquery.SourceFormat.CSV,
        skip_leading_rows=1,
        autodetect=False,
        schema=schema,
        write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
        allow_quoted_newlines=True,
        field_delimiter=",",
        encoding="UTF-8",
        max_bad_records=0,
    )

    gv_LOG.info("Starting BigQuery load from %s to %s", gcs_uri, table_id)
    result = client.load_table_from_uri(gcs_uri, table_id, job_config=job_config).result()
    gv_LOG.info("BigQuery load complete: %d rows loaded to %s", result.output_rows, table_id)

# ================================= Entrypoint =================================
def main() -> None:
    lv_parser = argparse.ArgumentParser(description="Territory Checks weekly ETL", add_help=True)
    lv_parser.add_argument(
        "--week_end",
        dest="week_end",
        type=str,
        required=False,
        help="Optional override for week end date in UTC, format YYYY-MM-DD (expected Sunday).",
    )

    lv_args, lv_unknown = lv_parser.parse_known_args()
    if lv_unknown:
        gv_LOG.info("Ignoring unknown CLI args (likely from platform): %s", lv_unknown)

    if lv_args.week_end:
        gv_LOG.info("Starting Territory Checks ETL job for explicit week_end=%s", lv_args.week_end)
    else:
        gv_LOG.info("Starting Territory Checks ETL job for last completed week (no override).")

    result = run(lv_override_week_end_str=lv_args.week_end)
    gv_LOG.info("Job finished.")
    gv_LOG.info("Excel local: %s", result.get("excel_local_path"))
    gv_LOG.info("Excel GCS  : %s", result.get("excel_gcs_uri"))
    gv_LOG.info("Raw local  : %s", result.get("raw_local_path"))
    gv_LOG.info("Raw GCS    : %s", result.get("raw_gcs_uri"))

if __name__ == "__main__":
    main()
